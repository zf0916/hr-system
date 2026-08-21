"""Loading the employee list from the spreadsheet HR sends.

The list does not exist yet and nobody has seen its columns. So the importer is
told, in an explicit mapping file, which sheet to read, which row the data
starts on, and which column letter holds which field. **Header text is never
matched on.** It is read back and echoed in the report, so a person can see at
a glance that `employee_number` is pointing at the column headed `No.` and not
at the one headed `Dept` — but nothing in the code decides anything from it.

Everything is checked before anything is written, and one bad row writes
nothing at all. A wrong mapping has to fail out loud; the failure this design
exists to prevent is a quiet load of plausible-looking wrong employees.

What it never does: pad or strip an employee number on the way in (SPEC §13).
The number is stored exactly as the sheet gave it. Padding belongs to the
separate matching key, which is rebuildable — `hr employees rekey`.
"""

from __future__ import annotations

import datetime as dt
import hashlib
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete, func, select

from app.models import (
    DailyAttendance,
    GatePass,
    LeaveRecord,
    DeviceUserMap,
    Employee,
    EmployeeAssignment,
    EmployeeGroup,
    EmployeeImport,
    EmployeeNumberKey,
    EmployeeNumberRule,
    EmploymentPeriod,
    Role,
    Section,
)
from app.xlsx_mapping import (
    Mapping,
    MappingError,
    RowProblem,
    cell_text,
    iter_rows,
    open_sheet,
    read_date,
    read_headers,
    read_mapping as read_mapping_file,
)

REQUIRED_COLUMNS = ("employee_number", "name", "section", "role", "group", "active_from")
OPTIONAL_COLUMNS = ("left_on", "device_pin")
ALL_COLUMNS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS
@dataclass
class StagedRow:
    row: int
    employee_number: str
    number_key: str
    number_from_numeric_cell: bool
    odd_number: bool
    name: str
    section: str
    role: str
    group: str
    active_from: dt.date
    left_on: dt.date | None
    device_pin: str | None
    # Why this row has no PIN, when it has none, and what was unusual about it
    # when it has one. A count of mappings written cannot tell a correctly
    # skipped blank from a rejected value, and those are different facts about
    # HR's list (SPEC §2).
    pin_note: str = ""


log = logging.getLogger("hr.employee_import")


@dataclass
class Result:
    problems: list[RowProblem] = field(default_factory=list)
    staged: list[StagedRow] = field(default_factory=list)
    blank_rows: list[int] = field(default_factory=list)
    new_vocabulary: list[tuple[str, str]] = field(default_factory=list)
    headers: dict[str, str] = field(default_factory=dict)
    written: dict[str, int] = field(default_factory=dict)


# ---- the rules, which are rows ---------------------------------------------


def number_rules(session) -> dict[str, str]:
    rows = session.scalars(select(EmployeeNumberRule)).all()
    rules = {r.key: r.value for r in rows}
    for required in ("expected_shape", "key_width", "key_pad"):
        if required not in rules:
            raise MappingError(
                f"employee_number_rule is missing {required!r}. Run `hr seed`."
            )
    return rules


def build_key(employee_number: str, rules: dict[str, str]) -> str:
    """Where padding is allowed to happen, and the only place (SPEC §2)."""
    width = int(rules["key_width"])
    pad = rules["key_pad"]
    return employee_number.strip().rjust(width, pad)


# ---- reading and checking --------------------------------------------------


def read_rows(workbook, mapping: Mapping, session, rules: dict[str, str],
              accept_odd_numbers: bool,
              accept_leading_zero_pins: bool = False) -> Result:
    result = Result()
    sheet = open_sheet(workbook, mapping)
    result.headers = read_headers(sheet, mapping)
    expected_shape = rules["expected_shape"]
    seen_numbers: dict[str, int] = {}
    seen_keys: dict[str, int] = {}
    seen_pins: dict[str, int] = {}

    known_sections = set(session.scalars(select(Section.code)))
    known_roles = set(session.scalars(select(Role.code)))
    known_groups = set(session.scalars(select(EmployeeGroup.code)))
    vocabulary = {
        "section": (known_sections, set()),
        "role": (known_roles, set()),
        "group": (known_groups, set()),
    }

    for row, cells, blank in iter_rows(sheet, mapping):
        if blank:
            result.blank_rows.append(row)
            continue

        problems_before = len(result.problems)

        # The number, exactly as given.
        raw_number = cells["employee_number"]
        number = cell_text(raw_number)
        from_numeric = isinstance(raw_number, (int, float)) and not isinstance(
            raw_number, bool
        )
        odd = not re.fullmatch(expected_shape, number)
        if not number.strip():
            result.problems.append(RowProblem(row, "employee_number", "is empty"))
        elif odd and not accept_odd_numbers:
            result.problems.append(RowProblem(
                row, "employee_number",
                f"{number!r} does not match the expected shape {expected_shape} "
                f"(employee_number_rule). It would be stored exactly as it is and "
                f"matched as {build_key(number, rules)!r} — pass "
                "--accept-odd-numbers if that is what you want",
            ))
        if number in seen_numbers:
            result.problems.append(RowProblem(
                row, "employee_number",
                f"{number!r} is already used on row {seen_numbers[number]}",
            ))
        key = build_key(number, rules) if number.strip() else ""
        if key and key in seen_keys and seen_keys[key] != row:
            result.problems.append(RowProblem(
                row, "employee_number",
                f"{number!r} matches as {key!r}, the same as row {seen_keys[key]} — "
                "two numbers that pad to one key are one person",
            ))
        if number.strip():
            seen_numbers.setdefault(number, row)
            seen_keys.setdefault(key, row)

        name = cell_text(cells["name"]).strip()
        if not name:
            result.problems.append(RowProblem(row, "name", "is empty"))

        values = {}
        for field_name in ("section", "role", "group"):
            value = cell_text(cells[field_name]).strip()
            values[field_name] = value
            if not value:
                result.problems.append(RowProblem(row, field_name, "is empty"))
                continue
            known, added = vocabulary[field_name]
            if value not in known and value not in added:
                added.add(value)

        active_from = read_date(
            cells["active_from"], mapping.date_format, "active_from", row,
            result.problems,
        )
        if active_from is None and not any(
            p.row == row and p.field == "active_from" for p in result.problems
        ):
            result.problems.append(RowProblem(row, "active_from", "is empty"))

        left_on = None
        if "left_on" in mapping.columns:
            left_on = read_date(
                cells["left_on"], mapping.date_format, "left_on", row,
                result.problems,
            )
        if active_from and left_on and left_on < active_from:
            result.problems.append(RowProblem(
                row, "left_on",
                f"{left_on.isoformat()} is before active_from "
                f"{active_from.isoformat()}",
            ))

        pin = None
        pin_note = "no device_pin column in the mapping"
        if "device_pin" in mapping.columns:
            pin = cell_text(cells["device_pin"]).strip() or None
            pin_note = "the Device ID cell is empty" if pin is None else ""

            # **The device refuses a leading zero in a user ID** (SPEC §2, §10,
            # observed at enrollment). A PIN like `0142` cannot exist on the
            # device, so a mapping row carrying one can never match a punch: it
            # looks like a working link and silently is not, and the employee's
            # punches go unattributed with nothing on screen to say why.
            if pin and len(pin) > 1 and pin.startswith("0"):
                if accept_leading_zero_pins:
                    pin_note = (
                        "accepted deliberately, and no punch from this device "
                        "can ever carry it (SPEC §2, §10)"
                    )
                else:
                    result.problems.append(RowProblem(
                        row, "device_pin",
                        f"{pin!r} starts with a leading zero, and the device "
                        "refuses one in a user ID (SPEC §10). A punch can never "
                        "carry this PIN, so the mapping would never match. "
                        "Correct the list, or re-run with "
                        "--accept-leading-zero-pins if this list belongs to "
                        "another device",
                    ))

            if pin and pin in seen_pins:
                result.problems.append(RowProblem(
                    row, "device_pin",
                    f"{pin!r} is already used on row {seen_pins[pin]} — one PIN "
                    "cannot belong to two employees",
                ))
            if pin:
                seen_pins.setdefault(pin, row)

        if len(result.problems) == problems_before and active_from:
            result.staged.append(StagedRow(
                row=row,
                employee_number=number,
                number_key=key,
                number_from_numeric_cell=from_numeric,
                odd_number=odd,
                name=name,
                section=values["section"],
                role=values["role"],
                group=values["group"],
                active_from=active_from,
                left_on=left_on,
                device_pin=pin,
                pin_note=pin_note,
            ))

    for kind, (_known, added) in vocabulary.items():
        for value in sorted(added):
            result.new_vocabulary.append((kind, value))
    return result


def check_against_database(session, result: Result) -> None:
    numbers = [s.employee_number for s in result.staged]
    if numbers:
        existing = set(
            session.scalars(
                select(Employee.employee_number).where(
                    Employee.employee_number.in_(numbers)
                )
            )
        )
        for staged in result.staged:
            if staged.employee_number in existing:
                result.problems.append(RowProblem(
                    staged.row, "employee_number",
                    f"{staged.employee_number!r} is already loaded. "
                    "--replace reloads the list from scratch",
                ))
    keys = [s.number_key for s in result.staged]
    if keys:
        taken = set(
            session.scalars(
                select(EmployeeNumberKey.key).where(EmployeeNumberKey.key.in_(keys))
            )
        )
        for staged in result.staged:
            if staged.number_key in taken:
                result.problems.append(RowProblem(
                    staged.row, "employee_number",
                    f"matching key {staged.number_key!r} is already taken",
                ))


# ---- writing ---------------------------------------------------------------


def clear_employees(session) -> dict[str, int]:
    """Everything the importer writes, and nothing else. Captured punches are
    untouched: nothing at capture resolves a PIN to an employee, so they do not
    depend on any of this.

    Daily attendance goes with it. Those rows are derived from punches, the
    mapping and the schedule (SPEC §3, layer 3), so reloading the list makes
    them stale rather than wrong — and they are rebuilt by
    `hr attendance build`, not re-collected. The punches they were built from
    are in the raw layer either way.
    """
    counts = {}
    for model in (
        DailyAttendance,
        DeviceUserMap,
        EmployeeAssignment,
        EmploymentPeriod,
        EmployeeNumberKey,
        Employee,
        EmployeeImport,
    ):
        counts[model.__tablename__] = session.execute(delete(model)).rowcount
    if counts.get("daily_attendance"):
        log.info(
            "%s daily attendance rows cleared with the list; rebuild them with "
            "`hr attendance build`", counts["daily_attendance"]
        )
    return counts


def write(session, mapping: Mapping, result: Result, source: Path,
          allow_new: set[str]) -> None:
    batch = EmployeeImport(
        source_filename=source.name,
        source_sha256=hashlib.sha256(source.read_bytes()).hexdigest(),
        mapping_filename=mapping.filename,
        mapping_text=mapping.text,
        row_count=len(result.staged),
    )
    session.add(batch)
    session.flush()

    models = {"section": Section, "role": Role, "group": EmployeeGroup}
    for kind, value in result.new_vocabulary:
        if kind in allow_new:
            session.add(models[kind](
                code=value, label=value, note=f"added by import of {source.name}"
            ))
    session.flush()

    written = {
        "employee": 0,
        "employee_number_key": 0,
        "employee_assignment": 0,
        "employment_period": 0,
        "device_user_map": 0,
    }
    for staged in result.staged:
        employee = Employee(
            employee_number=staged.employee_number, imported_from_id=batch.id
        )
        session.add(employee)
        session.flush()
        written["employee"] += 1

        session.add(EmployeeNumberKey(
            employee_id=employee.id,
            key=staged.number_key,
            built_by="employee_number_rule",
            note=None if not staged.odd_number else "built from an odd number",
        ))
        written["employee_number_key"] += 1

        session.add(EmployeeAssignment(
            employee_id=employee.id,
            effective_from=staged.active_from,
            effective_to=staged.left_on,
            name=staged.name,
            section_code=staged.section,
            role_code=staged.role,
            group_code=staged.group,
            imported_from_id=batch.id,
        ))
        written["employee_assignment"] += 1

        session.add(EmploymentPeriod(
            employee_id=employee.id,
            active_from=staged.active_from,
            left_on=staged.left_on,
            imported_from_id=batch.id,
        ))
        written["employment_period"] += 1

        if staged.device_pin:
            session.add(DeviceUserMap(
                serial_number=mapping.extra.get("device_serial"),
                pin=staged.device_pin,
                employee_id=employee.id,
                effective_from=staged.active_from,
                effective_to=staged.left_on,
                source=f"import of {source.name}",
                imported_from_id=batch.id,
            ))
            written["device_user_map"] += 1
    result.written = written


def run_import(session, source: Path, mapping_path: Path, *, replace: bool,
               allow_new: set[str], accept_odd_numbers: bool,
               accept_leading_zero_pins: bool = False) -> Result:
    """Read, check everything, then write — or write nothing at all."""
    mapping = read_mapping_file(
        mapping_path, REQUIRED_COLUMNS, OPTIONAL_COLUMNS, ("device_serial",)
    )
    rules = number_rules(session)

    workbook = load_workbook(source, data_only=True, read_only=False)
    try:
        result = read_rows(workbook, mapping, session, rules, accept_odd_numbers,
                           accept_leading_zero_pins)
    finally:
        workbook.close()

    if replace:
        # **HR entry is not derivable and is never cleared with the list.**
        # Daily attendance goes because it is rebuilt from punches; a leave
        # record and a gate pass are forms somebody typed off paper, and
        # deleting the employee they hang on would take them with it. Once HR
        # has typed anything, the list is corrected in place rather than
        # replaced wholesale (SPEC §5, §6).
        typed = {
            "leave_record": session.scalar(
                select(func.count()).select_from(LeaveRecord)) or 0,
            "gate_pass": session.scalar(
                select(func.count()).select_from(GatePass)) or 0,
        }
        if any(typed.values()):
            counted = ", ".join(f"{count} {name}" for name, count in typed.items()
                                if count)
            result.problems.append(RowProblem(
                None, "--replace",
                f"{counted} recorded against the employees now loaded. "
                "--replace deletes employees, and HR typed those forms off "
                "paper — they cannot be rebuilt from anything. Load without "
                "--replace, or decide what should happen to them first",
            ))
            return result
        result.written = {}
        cleared = clear_employees(session)
        session.flush()
        result.written["cleared"] = sum(cleared.values())

    check_against_database(session, result)

    # A new value is allowed only for the kind the operator named. Blanket
    # permission would disarm the best wrong-mapping detector there is: a
    # section column pointed at the names would quietly become eight sections.
    for kind, value in result.new_vocabulary:
        if kind not in allow_new:
            result.problems.append(RowProblem(
                0, kind,
                f"{value!r} is not a known {kind}. If this list really does "
                f"introduce {kind}s, pass --allow-new {kind} — and check first "
                "that the column is pointing where you meant",
            ))

    if result.problems:
        session.rollback()
        return result

    write(session, mapping, result, source, allow_new)
    return result
