"""Step 7: the Daily Workers Attendance sheet, in HR's existing layout.

**One render, three outputs.** `render` builds the sheet once, from the daily
attendance rows, and every emitter draws that same object: `to_text` for a
terminal, `to_json` for the browser — the screen, which is the system — and
`to_excel` for the file, which is the record HR files (SPEC §7). Every mark a
reader can see — the tick, the out-of-schedule time, the manual asterisk,
whether a column is shaded — is decided here and carried on the cell. **The
emitters choose fonts and column widths and nothing else**, which is what makes
it impossible for the file and the screen to disagree about what a day says.

What a cell holds (SPEC §7): a tick when the punch is on schedule, the actual
punch time when it is outside it, or a leave code. **Leave takes the cell when
there is leave** (SPEC §9 A49): a day with a leave record shows its sheet code,
and a day whose leave has no sheet code shows nothing rather than inventing a
letter — four of the seven form types have no code at all (§6).

Rest days and public holidays shade whole columns, from the calendar, never per
employee (SPEC §4). Manual punches are marked (SPEC §3).

Nothing here totals anything. Every period total is a query over the daily rows
and belongs to Milestone 3 (SPEC §3).
"""

from __future__ import annotations

import datetime as dt
from dataclasses import dataclass, field

from sqlalchemy import select

from app.hr_entry import leave_by_day, leave_codes
from app.models import (
    DailyAttendance,
    Employee,
    EmployeeAssignment,
    GroupSchedule,
    SheetSetting,
)
from app.schedule import effective_holiday, is_rest_day, schedule_for

# What a cell is showing, so a reader never has to infer it from the text.
TICK = "tick"
TIME = "time"
LEAVE = "leave"
EMPTY = "empty"


@dataclass
class Cell:
    employee_id: int
    date: dt.date
    text: str
    kind: str
    manual: bool = False
    leave_code: str | None = None
    punch_count: int = 0
    late_minutes: int | None = None
    provisional: bool = False
    detail: str = ""


@dataclass
class Column:
    date: dt.date
    day: int
    weekday: str
    shaded: bool
    shade_reason: str
    holiday_name: str | None = None
    rest_for: tuple[str, ...] = ()
    provisional_holiday: bool = False


@dataclass
class Row:
    employee_id: int
    employee_number: str
    name: str
    section_code: str
    role_code: str
    group_code: str
    page: int


@dataclass
class Sheet:
    title: str
    period_start: dt.date
    period_end: dt.date
    note_top_left: str
    note_is_unread: bool
    rows_per_page: int
    page_count: int
    headcount: int
    columns: list[Column]
    rows: list[Row]
    cells: dict[tuple[int, dt.date], Cell]
    legend: list[tuple[str, str]]
    notes: list[str] = field(default_factory=list)
    # How many cells say something that rests on a schedule row HR has never
    # confirmed. Counted at render, so the screen, the file and the text output
    # all quote the same number rather than each counting for themselves.
    provisional_cells: int = 0

    def cell(self, employee_id: int, day: dt.date) -> Cell:
        return self.cells[(employee_id, day)]

    def page(self, number: int) -> list[Row]:
        return [row for row in self.rows if row.page == number]


def settings(session) -> dict[str, str]:
    return {row.key: row.value for row in session.scalars(select(SheetSetting))}


def period_for(session, month: str) -> tuple[dt.date, dt.date]:
    """The period one sheet covers, from the rule row (A40).

    `calendar_month` is the assumption. The 10th/15th/20th cut-offs on the paper
    sheet may be deadlines rather than boundaries (SPEC §5), and until that is
    answered this is a row that an UPDATE corrects.
    """
    rule = settings(session).get("sheet.period_rule", "calendar_month")
    year, mon = (int(part) for part in month.split("-"))
    start = dt.date(year, mon, 1)
    if rule != "calendar_month":
        raise ValueError(
            f"sheet.period_rule is {rule!r}, and only 'calendar_month' is "
            "implemented. The rule is a row; teach the renderer the new one "
            "before changing it"
        )
    end_month = start.replace(day=28) + dt.timedelta(days=4)
    end = end_month - dt.timedelta(days=end_month.day)
    return start, end


def resolve_period(session, month: str | None = None, start=None,
                   end=None) -> tuple[dt.date, dt.date]:
    """A month name, or two dates, into the period one render covers.

    One resolver for every caller — the terminal, the browser and the download
    — so that "August" cannot mean one span on the screen and another in the
    file. Dates may be `date` objects or `YYYY-MM-DD` strings.
    """
    if month:
        return period_for(session, month)
    if not (start and end):
        raise ValueError("give a month as YYYY-MM, or both a start and an end")

    def as_date(value):
        if isinstance(value, dt.date):
            return value
        return dt.datetime.strptime(value, "%Y-%m-%d").date()

    return as_date(start), as_date(end)


def _cell_for(row: DailyAttendance | None, employee_id: int, day: dt.date,
              marks: dict[str, str], leave=None) -> Cell:
    """One day for one employee, from its daily row and nothing else.

    A38: a tick when every punch that day is inside the schedule; otherwise the
    punch times that fall outside it — the first in when it is later than the
    scheduled start plus grace, the last out when it is earlier than the
    scheduled end. Both, when both are.
    """
    # **A coded leave day is a leave cell** (A49). A day somebody was on leave
    # is not a day they failed to punch, and the code is what HR writes.
    leave_detail = ""
    if leave is not None:
        leave_detail = f"leave record {leave.record_id}"
        if leave.type_label:
            leave_detail += f": {leave.type_label}"
        leave_detail += f", {leave.days} day(s)"
        if row is not None and row.punch_count:
            leave_detail += f"; {row.punch_count} punch(es) on the same day"

        if leave.sheet_code:
            return Cell(
                employee_id, day, leave.sheet_code, LEAVE,
                leave_code=leave.sheet_code,
                punch_count=row.punch_count if row is not None else 0,
                provisional=row.schedule_provisional if row is not None else False,
                detail=leave_detail,
            )
        # **No code on the record: the cell falls through to the punches.**
        # Four of the seven form types have no legend letter (§6) and none is
        # borrowed from the type — but blanking the cell would hide a punch
        # that really happened, so what the punches say is what shows.
        leave_detail += " — the record carries no sheet code, so this cell "
        leave_detail += "shows the punches instead of a letter"

    if row is None:
        return Cell(employee_id, day, "", EMPTY,
                    detail="no daily row: not employed, or not built")

    leave_code = None

    if row.punch_count == 0:
        return Cell(employee_id, day, "", EMPTY, punch_count=0,
                    provisional=row.schedule_provisional,
                    detail=leave_detail
                    or "no punch — a fact, not an absence (SPEC §3)")

    time_format = marks.get("sheet.time_format", "%H:%M")
    outside: list[str] = []
    if row.scheduled_start is None:
        # Nothing to be inside of: a rest day, a closed holiday, or no schedule.
        # The punch is real and shows as a time.
        if row.first_in:
            outside.append(row.first_in.strftime(time_format))
        if row.last_out:
            outside.append(row.last_out.strftime(time_format))
    else:
        if row.late_minutes:
            outside.append(row.first_in.strftime(time_format))
        if (row.last_out is not None and row.scheduled_end is not None
                and row.last_out < row.scheduled_end):
            outside.append(row.last_out.strftime(time_format))

    manual = bool(row.first_in_manual or row.last_out_manual)
    mark = marks.get("sheet.mark_manual", "*") if manual else ""

    if outside:
        text, kind = "/".join(outside), TIME
    else:
        text, kind = marks.get("sheet.mark_on_schedule", "✓"), TICK

    detail = f"{row.punch_count} punches"
    if leave_detail:
        detail = f"{leave_detail}; {detail}"
    if row.manual_punch_count:
        detail += f", {row.manual_punch_count} entered by a person"
    if row.duplicate_pushes:
        detail += f", {row.duplicate_pushes} re-pushes dropped"

    return Cell(
        employee_id, day, text + mark, kind, manual=manual,
        leave_code=leave_code, punch_count=row.punch_count,
        late_minutes=row.late_minutes, provisional=row.schedule_provisional,
        detail=detail,
    )


def render(session, start: dt.date, end: dt.date,
           section_code: str | None = None) -> Sheet:
    """Build the sheet once. Both outputs draw this."""
    if end < start:
        raise ValueError(f"{end} is before {start}")

    marks = settings(session)
    rows_per_page = int(marks.get("sheet.rows_per_page", "30"))

    assignments = session.execute(
        select(EmployeeAssignment, Employee.employee_number)
        .join(Employee, Employee.id == EmployeeAssignment.employee_id)
        .where(
            EmployeeAssignment.effective_from <= end,
            (EmployeeAssignment.effective_to.is_(None))
            | (EmployeeAssignment.effective_to >= start),
        )
        .order_by(EmployeeAssignment.section_code, Employee.employee_number)
    ).all()
    if section_code:
        assignments = [a for a in assignments if a[0].section_code == section_code]

    sheet_rows: list[Row] = []
    for index, (assignment, number) in enumerate(assignments):
        sheet_rows.append(Row(
            employee_id=assignment.employee_id,
            employee_number=number,
            name=assignment.name,
            section_code=assignment.section_code,
            role_code=assignment.role_code,
            group_code=assignment.group_code,
            page=index // rows_per_page + 1,
        ))

    groups = sorted({row.group_code for row in sheet_rows})

    columns: list[Column] = []
    notes: list[str] = []
    day = start
    while day <= end:
        holiday = effective_holiday(session, day)
        closes = bool(holiday and holiday.closes)
        resting = []
        for group in groups:
            schedule = schedule_for(session, group, day)
            if schedule is not None and is_rest_day(schedule, day):
                resting.append(group)

        # A42: a column shades when the day is closed for everyone on the sheet.
        # Groups that rest on different days would break whole-column shading,
        # and that is a fact about the sheet rather than something to paper over.
        all_rest = bool(groups) and len(resting) == len(groups)
        shaded = closes or all_rest
        reason = ""
        if closes:
            reason = holiday.name
        elif all_rest:
            reason = "rest day"
        elif resting:
            reason = "rest day for some groups only — not shaded"
            notes.append(
                f"{day}: rest day for {', '.join(resting)} but not for "
                f"{', '.join(g for g in groups if g not in resting)}, so the "
                "column is not shaded (SPEC §9 A42)"
            )

        columns.append(Column(
            date=day, day=day.day, weekday=day.strftime("%a"), shaded=shaded,
            shade_reason=reason,
            holiday_name=holiday.name if holiday else None,
            rest_for=tuple(resting),
            provisional_holiday=bool(holiday and holiday.provisional),
        ))
        day += dt.timedelta(days=1)

    leave = leave_by_day(session, start, end)
    daily = {
        (row.employee_id, row.attendance_day): row
        for row in session.scalars(
            select(DailyAttendance).where(
                DailyAttendance.attendance_day >= start,
                DailyAttendance.attendance_day <= end,
            )
        )
    }
    cells = {
        (row.employee_id, column.date): _cell_for(
            daily.get((row.employee_id, column.date)), row.employee_id,
            column.date, marks,
            leave=leave.get((row.employee_id, column.date)))
        for row in sheet_rows
        for column in columns
    }

    # A cell resting on an unconfirmed schedule is only worth counting when it
    # says something the schedule decided. A blank day and a leave code say
    # nothing about a schedule; **a tick is the claim "this was on time"**, and
    # an out-of-schedule time is the claim "this was not".
    provisional_cells = sum(
        1 for cell in cells.values()
        if cell.provisional and cell.kind in (TICK, TIME)
    )
    if provisional_cells:
        notes.append(
            f"{provisional_cells} cell(s) rest on schedule rows HR has not "
            "confirmed: every tick and every out-of-schedule time on this "
            "sheet was decided against a provisional schedule. The punch times "
            "are real; whether they are late is arithmetic on a guess."
        )
    if any(column.provisional_holiday for column in columns):
        notes.append("Some holidays on this calendar are provisional.")
    coded = sum(1 for cell in cells.values() if cell.leave_code)
    uncoded = sum(1 for cell in cells.values()
                  if not cell.leave_code and "leave record" in cell.detail)
    if coded or uncoded:
        note = f"{coded} day(s) of leave show a code"
        if uncoded:
            note += (f", and {uncoded} day(s) of leave show no letter because "
                     "the record carries no sheet code — four form types have "
                     "none (SPEC §6). Those cells show whatever the punches "
                     "say")
        notes.append(note + ".")
    if not any(cell.punch_count for cell in cells.values()):
        notes.append("No punches fell in this period for anyone on the sheet.")

    note = marks.get("sheet.note_top_left", "")
    return Sheet(
        title=marks.get("sheet.title", "DAILY WORKERS ATTENDANCE"),
        period_start=start,
        period_end=end,
        note_top_left=note,
        note_is_unread=not note.strip(),
        rows_per_page=rows_per_page,
        page_count=max(1, (len(sheet_rows) + rows_per_page - 1) // rows_per_page),
        headcount=len(sheet_rows),
        columns=columns,
        rows=sheet_rows,
        cells=cells,
        legend=legend_rows(session),
        notes=notes,
        provisional_cells=provisional_cells,
    )


def legend_rows(session) -> list[tuple[str, str]]:
    """The sheet legend: the leave codes, and the sheet's own marks.

    **The codes come from `leave_code` rows**, not from this file — §13 says a
    leave code is a row rather than a constant, and HR's paper legend is what
    those rows were seeded from.
    """
    marks = settings(session)
    legend = [(row.code, row.label) for row in leave_codes(session)]
    legend += [
        (marks.get("sheet.mark_on_schedule", "✓"), "punch on schedule"),
        ("08:20", "the actual punch time, when it is outside the schedule"),
        (marks.get("sheet.mark_manual", "*"),
         "punch entered by a person, not the device"),
        ("(blank)", "no punch — a fact, never an absence (SPEC §3)"),
        ("(shaded)", "rest day or a public holiday the factory closes for"),
    ]
    return legend


# ---- the screen ----------------------------------------------------------


def to_text(sheet: Sheet, width: int | None = None) -> str:
    """The screen. Same object, same marks, same shading — as text.

    The column is as wide as the widest thing in it. A cell showing both an
    out-of-schedule arrival and an out-of-schedule departure is eleven
    characters, and a fixed width would push the grid out of line — which is a
    reader mistaking one day for another.
    """
    out: list[str] = []

    def column_width(column: Column) -> int:
        if width is not None:
            return width
        longest = max(
            (len(sheet.cell(row.employee_id, column.date).text)
             for row in sheet.rows),
            default=0,
        )
        return max(5, longest + 1)

    widths = [column_width(column) for column in sheet.columns]
    out.append(sheet.title)
    out.append(f"{sheet.period_start} → {sheet.period_end}"
               f"   headcount {sheet.headcount}"
               f"   {sheet.page_count} page(s) of {sheet.rows_per_page}")
    if sheet.note_is_unread:
        out.append("note (top-left): [ ]  <- not read; nothing is guessed here "
                   "(SPEC §9 A41)")
    else:
        out.append(f"note (top-left): {sheet.note_top_left}")

    header = " " * 26 + "".join(
        f"{c.day:>{w}}" for c, w in zip(sheet.columns, widths))
    weekdays = " " * 26 + "".join(
        f"{c.weekday[:2]:>{w}}" for c, w in zip(sheet.columns, widths))
    shading = " " * 26 + "".join(
        f"{'▓▓▓▓' if c.shaded else '':>{w}}"
        for c, w in zip(sheet.columns, widths))
    out += ["", header, weekdays, shading]

    section = None
    for row in sheet.rows:
        if row.section_code != section:
            section = row.section_code
            out.append(f"-- {section}")
        line = f"{row.employee_number:>6} {row.name[:14]:<14} {row.role_code[:3]:<4}"
        for column, w in zip(sheet.columns, widths):
            line += f"{sheet.cell(row.employee_id, column.date).text:>{w}}"
        out.append(line)

    out.append("")
    out.append("legend: " + "  ".join(f"{code} {label}"
                                      for code, label in sheet.legend))
    for note in sheet.notes:
        out.append(f"note: {note}")
    for column in sheet.columns:
        if column.shaded:
            out.append(f"shaded: {column.date} — {column.shade_reason}")
    return "\n".join(out)


# ---- the browser --------------------------------------------------------


def to_json(sheet: Sheet) -> dict:
    """The same object again, as plain data for the browser.

    **A third emitter, not a second render.** Everything a screen can show is
    carried on the Cell already — the text, what kind of thing it is, whether a
    person entered it, whether it rests on a provisional schedule — so this
    function reads fields and formats dates and does nothing else. The browser
    cannot compute a different answer from the file because it is never given
    the ingredients, only the answer (SPEC §7).

    Cells are keyed `"{employee_id}:{date}"`, because JSON has no tuple.
    """
    return {
        "title": sheet.title,
        "period_start": sheet.period_start.isoformat(),
        "period_end": sheet.period_end.isoformat(),
        "note_top_left": sheet.note_top_left,
        "note_is_unread": sheet.note_is_unread,
        "rows_per_page": sheet.rows_per_page,
        "page_count": sheet.page_count,
        "headcount": sheet.headcount,
        "provisional_cells": sheet.provisional_cells,
        "columns": [
            {
                "date": column.date.isoformat(),
                "day": column.day,
                "weekday": column.weekday,
                "shaded": column.shaded,
                "shade_reason": column.shade_reason,
                "holiday_name": column.holiday_name,
                "rest_for": list(column.rest_for),
                "provisional_holiday": column.provisional_holiday,
            }
            for column in sheet.columns
        ],
        "rows": [
            {
                "employee_id": row.employee_id,
                "employee_number": row.employee_number,
                "name": row.name,
                "section_code": row.section_code,
                "role_code": row.role_code,
                "group_code": row.group_code,
                "page": row.page,
            }
            for row in sheet.rows
        ],
        "cells": {
            f"{employee_id}:{day.isoformat()}": {
                "text": cell.text,
                "kind": cell.kind,
                "manual": cell.manual,
                "leave_code": cell.leave_code,
                "punch_count": cell.punch_count,
                "late_minutes": cell.late_minutes,
                "provisional": cell.provisional,
                "detail": cell.detail,
            }
            for (employee_id, day), cell in sheet.cells.items()
        },
        "legend": [{"code": code, "label": label}
                   for code, label in sheet.legend],
        "notes": list(sheet.notes),
    }


# ---- the file -----------------------------------------------------------


def to_excel(sheet: Sheet, path) -> None:
    """Write the record to a file. The bytes come from `to_bytes`."""
    from pathlib import Path as _Path

    _Path(path).write_bytes(to_bytes(sheet))


def to_bytes(sheet: Sheet) -> bytes:
    """The record, as bytes. **The same period always produces the same file.**

    One function makes the file, so the browser's download and `hr sheet
    export` cannot be two different files that merely look alike — the download
    is these bytes, unaltered.

    Byte-for-byte reproducibility is not decoration. The Excel file is the
    filed record (SPEC §7), and a filed record you cannot compare against a
    fresh render is a record you have to take on trust. Two things would
    otherwise vary with the clock: the archive's per-member timestamps, and the
    created/modified stamps openpyxl writes into the document properties. Both
    are pinned below — **the file is stamped with the period it covers, not
    with the moment somebody exported it**, which is also the truer statement
    of what the file is.
    """
    import io
    import zipfile

    from openpyxl.writer.excel import ExcelWriter

    workbook = _build_workbook(sheet)
    # The document's own timestamps describe the period, never the export.
    stamp = dt.datetime.combine(sheet.period_end, dt.time(0, 0))
    workbook.properties.created = stamp
    workbook.properties.modified = stamp

    # `workbook.save()` overwrites `modified` with the clock on its way past,
    # so the archive is written through openpyxl's writer instead — the same
    # writer `save()` uses, with the one line that reaches for the time left
    # out.
    buffer = io.BytesIO()
    ExcelWriter(
        workbook,
        zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED, allowZip64=True),
    ).save()

    # Each archive member is also stamped with the current local time. Rewrite
    # them at a fixed instant, in the order openpyxl produced them.
    original = zipfile.ZipFile(io.BytesIO(buffer.getvalue()))
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as archive:
        for member in original.infolist():
            fixed = zipfile.ZipInfo(member.filename, date_time=(1980, 1, 1, 0, 0, 0))
            fixed.compress_type = member.compress_type
            fixed.external_attr = member.external_attr
            archive.writestr(fixed, original.read(member.filename))
    return out.getvalue()


def _build_workbook(sheet: Sheet):
    """The record. The same object as the screen, drawn into HR's layout.

    Every string written here comes off the Sheet. Fonts, widths and fills are
    this function's business; **what a cell says is not** (SPEC §7).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.pagebreak import Break
    from openpyxl.worksheet.properties import PageSetupProperties

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Attendance"

    grey = PatternFill("solid", fgColor="D9D9D9")
    manual_fill = PatternFill("solid", fgColor="FFF2CC")
    unread = PatternFill("solid", fgColor="FCE4D6")
    thin = Side(style="thin", color="999999")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    centre = Alignment(horizontal="center", vertical="center")

    worksheet["A1"] = sheet.title
    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet["A2"] = f"{sheet.period_start} to {sheet.period_end}"
    worksheet["A3"] = f"headcount {sheet.headcount}"

    # The top-left note: the cell is empty, and the cell beside it says why.
    worksheet["A4"] = sheet.note_top_left
    if sheet.note_is_unread:
        worksheet["A4"].fill = unread
        worksheet["B4"] = "note not read — nothing guessed (SPEC §9 A41)"
        worksheet["B4"].font = Font(italic=True, color="C00000")

    header_row = 6
    worksheet.cell(header_row, 1, "No.").font = Font(bold=True)
    worksheet.cell(header_row, 2, "Name").font = Font(bold=True)
    worksheet.cell(header_row, 3, "Section").font = Font(bold=True)
    worksheet.cell(header_row, 4, "Role").font = Font(bold=True)
    for index, column in enumerate(sheet.columns):
        cell = worksheet.cell(header_row, 5 + index, column.day)
        cell.font = Font(bold=True)
        cell.alignment = centre
        cell.border = box
        weekday = worksheet.cell(header_row + 1, 5 + index, column.weekday)
        weekday.alignment = centre
        weekday.border = box
        if column.shaded:
            cell.fill = grey
            weekday.fill = grey

    first_data_row = header_row + 2
    for offset, row in enumerate(sheet.rows):
        excel_row = first_data_row + offset
        worksheet.cell(excel_row, 1, row.employee_number)
        worksheet.cell(excel_row, 2, row.name)
        worksheet.cell(excel_row, 3, row.section_code)
        worksheet.cell(excel_row, 4, row.role_code)
        for index, column in enumerate(sheet.columns):
            model_cell = sheet.cell(row.employee_id, column.date)
            cell = worksheet.cell(excel_row, 5 + index, model_cell.text or None)
            cell.alignment = centre
            cell.border = box
            if column.shaded:
                cell.fill = grey
            elif model_cell.manual:
                cell.fill = manual_fill

    # How it prints. The file is the filed record, not a screen dump.
    page = page_layout(sheet)
    worksheet.page_setup.orientation = page["orientation"]
    worksheet.page_setup.fitToWidth = page["fit_to_width"]
    worksheet.page_setup.fitToHeight = page["fit_to_height"]
    worksheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    # The employee header and the weekday row repeat on every page: a page of
    # ticks with no day numbers above them cannot be read at all.
    worksheet.print_title_rows = page["print_title_rows"]
    for row_number in page["row_breaks"]:
        worksheet.row_breaks.append(Break(id=row_number))

    note_row = page["legend_row"]
    worksheet.cell(note_row, 1, "Legend").font = Font(bold=True)
    for index, (code, label) in enumerate(sheet.legend, start=1):
        worksheet.cell(note_row + index, 1, code)
        worksheet.cell(note_row + index, 2, label)
    for index, note in enumerate(sheet.notes, start=len(sheet.legend) + 2):
        worksheet.cell(note_row + index, 1, note)

    worksheet.column_dimensions["A"].width = 8
    worksheet.column_dimensions["B"].width = 26
    worksheet.column_dimensions["C"].width = 14
    worksheet.column_dimensions["D"].width = 18
    worksheet.freeze_panes = worksheet.cell(first_data_row, 5)
    last_column = 4 + len(sheet.columns)
    worksheet.print_area = (
        f"A1:{worksheet.cell(1, last_column).column_letter}"
        f"{note_row + len(sheet.legend) + len(sheet.notes) + 2}"
    )
    return workbook


def page_layout(sheet: Sheet) -> dict:
    """How the file prints, derived from the sheet and its rows-per-page row.

    **The Excel file is the record HR files (SPEC §7), so it is a printed
    artefact** — a spreadsheet that prints its day columns across four pages
    with no header on three of them is not the sheet HR keeps, whatever the
    cells say. Everything here follows from `sheet.rows_per_page`, which is a
    row (§9 A39), so changing the page length is an UPDATE.

    One function, used by the writer and by the check that reads the file back,
    so the two cannot drift apart silently.
    """
    layout = excel_layout()
    first = layout["first_data_row"]
    breaks = []
    # A break after every full page of employee rows, but never after the last
    # one — that would print a blank page.
    for offset in range(sheet.rows_per_page, len(sheet.rows), sheet.rows_per_page):
        breaks.append(first + offset - 1)
    legend_row = first + len(sheet.rows) + 2
    # The legend and the notes start their own page. Mid-page they read as a
    # footnote to whichever employees happen to be above them.
    if sheet.rows:
        breaks.append(legend_row - 1)
    return {
        "orientation": "landscape",
        "fit_to_width": 1,
        "fit_to_height": 0,
        "print_title_rows": f"{layout['header_row']}:{layout['weekday_row']}",
        "row_breaks": sorted(set(breaks)),
        "legend_row": legend_row,
        "rows_per_page": sheet.rows_per_page,
    }


def excel_layout() -> dict:
    """Where `to_excel` puts things, so a reader does not have to guess.

    Used by the readback check in `tools/`, which is a verification of the file
    this code wrote — never an ingest path. Nothing in the application reads a
    sheet back in (SPEC §7, §13).
    """
    return {
        "header_row": 6,
        "weekday_row": 7,
        "first_data_row": 8,
        "first_day_column": 5,
        "number_column": 1,
    }
