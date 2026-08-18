"""Loading a year of public holidays, and changing one date afterwards.

The 2026 list does not exist yet and its columns are HR's to fill in, so this
reads the same way the employee list does: an explicit mapping naming the sheet,
the rows and the column letters, with header text echoed but never matched on.

Two ways to change the calendar, and they do not fight:

  * a full re-upload replaces the uploaded rows for one year;
  * a per-date adjustment is a row in its own table, so it survives a re-upload
    and is applied on top of whatever the new upload says.

A re-upload therefore never silently discards a correction. It reports every
adjustment that still stands, and which of them the new upload has made
redundant. The alternative — dropping adjustments on re-upload — would lose a
decision someone made deliberately, without saying so.
"""

from __future__ import annotations

import datetime as dt
import hashlib
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete, extract, select

from app.models import Holiday, HolidayAdjustment, HolidayScope, HolidayUpload
from app.xlsx_mapping import (
    MappingError,
    RowProblem,
    cell_text,
    iter_rows,
    open_sheet,
    read_date,
    read_headers,
    read_mapping,
)

REQUIRED_COLUMNS = ("date", "name", "scope", "closes")
OPTIONAL_COLUMNS = ("note",)

TRUE_TOKENS = {"yes", "y", "true", "t", "1", "close", "closed"}
FALSE_TOKENS = {"no", "n", "false", "f", "0", "open", "work", "working"}


@dataclass
class StagedHoliday:
    row: int
    date: dt.date
    name: str
    scope: str
    closes: bool
    note: str | None


@dataclass
class Result:
    problems: list[RowProblem] = field(default_factory=list)
    staged: list[StagedHoliday] = field(default_factory=list)
    blank_rows: list[int] = field(default_factory=list)
    new_scopes: list[str] = field(default_factory=list)
    headers: dict[str, str] = field(default_factory=dict)
    replaced: int = 0
    written: int = 0
    adjustments: list[tuple[dt.date, str, str]] = field(default_factory=list)


def read_closes(value, row: int, problems: list[RowProblem]) -> bool | None:
    text = cell_text(value).strip().lower()
    if not text:
        problems.append(RowProblem(
            row, "closes",
            "is empty — say whether the company actually closes that day. A "
            "gazetted holiday the factory works is a real case",
        ))
        return None
    if text in TRUE_TOKENS:
        return True
    if text in FALSE_TOKENS:
        return False
    problems.append(RowProblem(
        row, "closes",
        f"{cell_text(value)!r} is not yes or no (accepted: "
        f"{sorted(TRUE_TOKENS)} / {sorted(FALSE_TOKENS)})",
    ))
    return None


def read_rows(workbook, mapping, session, year: int) -> Result:
    result = Result()
    sheet = open_sheet(workbook, mapping)
    result.headers = read_headers(sheet, mapping)

    known_scopes = set(session.scalars(select(HolidayScope.code)))
    seen_dates: dict[dt.date, int] = {}
    new_scopes: set[str] = set()

    for row, cells, blank in iter_rows(sheet, mapping):
        if blank:
            result.blank_rows.append(row)
            continue

        problems_before = len(result.problems)

        date = read_date(cells["date"], mapping.date_format, "date", row,
                         result.problems)
        if date is None and len(result.problems) == problems_before:
            result.problems.append(RowProblem(row, "date", "is empty"))
        if date is not None:
            if date.year != year:
                result.problems.append(RowProblem(
                    row, "date",
                    f"{date.isoformat()} is not in {year} — this upload is for "
                    f"{year}, and a year is replaced whole",
                ))
            if date in seen_dates:
                result.problems.append(RowProblem(
                    row, "date",
                    f"{date.isoformat()} is already on row {seen_dates[date]}. "
                    "Two holidays on one date are one row with one name",
                ))
            else:
                seen_dates[date] = row

        name = cell_text(cells["name"]).strip()
        if not name:
            result.problems.append(RowProblem(row, "name", "is empty"))

        scope = cell_text(cells["scope"]).strip()
        if not scope:
            result.problems.append(RowProblem(row, "scope", "is empty"))
        elif scope not in known_scopes:
            new_scopes.add(scope)

        closes = read_closes(cells["closes"], row, result.problems)
        note = None
        if "note" in mapping.columns:
            note = cell_text(cells["note"]).strip() or None

        if len(result.problems) == problems_before and date and closes is not None:
            result.staged.append(StagedHoliday(
                row=row, date=date, name=name, scope=scope, closes=closes, note=note
            ))

    result.new_scopes = sorted(new_scopes)
    return result


def describe_adjustments(session, year: int) -> list[tuple[dt.date, str, str]]:
    """What each adjustment for that year does, now that the upload has changed.

    This is the report that keeps a re-upload from quietly overriding a
    decision, or from leaving a stale correction in place unnoticed.
    """
    from app.schedule import effective_holiday

    rows = session.scalars(
        select(HolidayAdjustment)
        .where(extract("year", HolidayAdjustment.holiday_date) == year)
        .order_by(HolidayAdjustment.holiday_date, HolidayAdjustment.made_at)
    ).all()

    latest: dict[dt.date, HolidayAdjustment] = {}
    for row in rows:
        latest[row.holiday_date] = row

    described = []
    for date, adjustment in sorted(latest.items()):
        uploaded = session.scalars(
            select(Holiday).where(Holiday.holiday_date == date)
        ).first()
        effective = effective_holiday(session, date)

        if adjustment.action == "remove":
            if uploaded is None:
                described.append((
                    date, "no longer changes anything",
                    "it removed a date the new upload does not have",
                ))
            else:
                described.append((
                    date, "still applies",
                    f"the upload has {uploaded.name!r}; the adjustment removes it "
                    f"({adjustment.reason})",
                ))
            continue

        if uploaded is None:
            described.append((
                date, "still applies",
                f"adds {effective.name!r}, which the upload does not have "
                f"({adjustment.reason})",
            ))
            continue

        differences = []
        if adjustment.closes is not None and adjustment.closes != uploaded.closes:
            differences.append(
                f"closes {uploaded.closes} -> {adjustment.closes}"
            )
        if adjustment.name and adjustment.name != uploaded.name:
            differences.append(f"name {uploaded.name!r} -> {adjustment.name!r}")
        if adjustment.scope_code and adjustment.scope_code != uploaded.scope_code:
            differences.append(
                f"scope {uploaded.scope_code!r} -> {adjustment.scope_code!r}"
            )
        if differences:
            described.append((
                date, "still applies",
                f"{'; '.join(differences)} ({adjustment.reason})",
            ))
        else:
            described.append((
                date, "no longer changes anything",
                "the new upload already says what the adjustment said",
            ))
    return described


def run_import(session, source: Path, mapping_path: Path, *, year: int,
               replace: bool, allow_new: set[str], provisional: bool) -> Result:
    mapping = read_mapping(mapping_path, REQUIRED_COLUMNS, OPTIONAL_COLUMNS)

    workbook = load_workbook(source, data_only=True, read_only=False)
    try:
        result = read_rows(workbook, mapping, session, year)
    finally:
        workbook.close()

    for scope in result.new_scopes:
        if "scope" not in allow_new:
            result.problems.append(RowProblem(
                0, "scope",
                f"{scope!r} is not a known holiday scope. Pass --allow-new scope "
                "if the list really introduces one",
            ))

    existing = session.scalars(
        select(Holiday).where(extract("year", Holiday.holiday_date) == year)
    ).all()
    if existing and not replace:
        result.problems.append(RowProblem(
            0, "year",
            f"{year} already has {len(existing)} holidays loaded. --replace "
            "replaces the year; per-date changes are `hr calendar adjust`, and "
            "those survive a replace",
        ))

    if result.problems:
        session.rollback()
        return result

    if replace:
        result.replaced = session.execute(
            delete(Holiday).where(extract("year", Holiday.holiday_date) == year)
        ).rowcount
        session.flush()

    for scope in result.new_scopes:
        session.add(HolidayScope(
            code=scope, label=scope, note=f"added by upload of {source.name}"
        ))
    session.flush()

    upload = HolidayUpload(
        year=year,
        source_filename=source.name,
        source_sha256=hashlib.sha256(source.read_bytes()).hexdigest(),
        mapping_filename=mapping.filename,
        mapping_text=mapping.text,
        row_count=len(result.staged),
        provisional=provisional,
    )
    session.add(upload)
    session.flush()

    for staged in result.staged:
        session.add(Holiday(
            holiday_date=staged.date,
            name=staged.name,
            scope_code=staged.scope,
            closes=staged.closes,
            provisional=provisional,
            upload_id=upload.id,
            note=staged.note,
        ))
    session.flush()
    result.written = len(result.staged)

    # Adjustments were never touched by the replace. Say what they do now.
    result.adjustments = describe_adjustments(session, year)
    return result


def adjust(session, *, date: dt.date, reason: str, made_by: str, remove: bool = False,
           closes: bool | None = None, name: str | None = None,
           scope: str | None = None) -> HolidayAdjustment:
    """One date changed, as a row beside the upload rather than an edit to it."""
    if remove and (closes is not None or name or scope):
        raise ValueError("--remove says the date is not a holiday; it takes no other values")
    if not remove and closes is None and not name and not scope:
        raise ValueError("nothing to change: give --closes, --name, --scope, or --remove")

    uploaded = session.scalars(
        select(Holiday).where(Holiday.holiday_date == date)
    ).first()
    if remove and uploaded is None:
        raise ValueError(f"{date.isoformat()} is not a holiday in the loaded calendar")
    if not remove and uploaded is None and not (name and scope and closes is not None):
        raise ValueError(
            f"{date.isoformat()} is not in the loaded calendar, so adding it needs "
            "--name, --scope and --closes"
        )
    if scope is not None and session.get(HolidayScope, scope) is None:
        raise ValueError(f"{scope!r} is not a known holiday scope")

    adjustment = HolidayAdjustment(
        holiday_date=date,
        action="remove" if remove else "set",
        name=name,
        scope_code=scope,
        closes=closes,
        reason=reason,
        made_by=made_by,
    )
    session.add(adjustment)
    session.flush()
    return adjustment
