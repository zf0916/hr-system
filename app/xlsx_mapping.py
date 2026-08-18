"""Reading a spreadsheet by an explicit mapping, never by its headers.

Shared by the employee importer and the holiday importer. Both are given files
nobody has seen the shape of — HR's employee list has not arrived, and the
holiday list is filled in by hand — so both are told which sheet, which rows and
which column letters, and neither ever matches on header text. Headers are read
only to be echoed back, so a person can see where a mapping is pointing.
"""

from __future__ import annotations

import datetime as dt
import re
import tomllib
from dataclasses import dataclass
from pathlib import Path

from openpyxl.utils import column_index_from_string, get_column_letter

BASE_KEYS = {
    "sheet",
    "sheet_index",
    "header_row",
    "first_data_row",
    "last_data_row",
    "date_format",
    "columns",
}


class MappingError(Exception):
    """The mapping file itself is wrong. Nothing is read."""


@dataclass
class RowProblem:
    row: int
    field: str
    message: str

    def __str__(self) -> str:
        where = f"row {self.row}: " if self.row else ""
        return f"{where}{self.field}: {self.message}"


@dataclass
class Mapping:
    sheet: str | None
    sheet_index: int | None
    header_row: int | None
    first_data_row: int
    last_data_row: int | None
    date_format: str | None
    columns: dict[str, int]  # field -> 1-based column index
    extra: dict
    text: str
    filename: str


def _column_index(field_name: str, value) -> int:
    """A column letter or a 1-based number. Never a header name."""
    if isinstance(value, bool):
        raise MappingError(f"columns.{field_name}: {value!r} is not a column")
    if isinstance(value, int):
        if value < 1:
            raise MappingError(f"columns.{field_name}: column numbers start at 1")
        return value
    if isinstance(value, str) and re.fullmatch(r"[A-Za-z]{1,3}", value.strip()):
        return column_index_from_string(value.strip().upper())
    raise MappingError(
        f"columns.{field_name}: {value!r} is not a column letter or number. "
        "Header names are never matched on — give the letter."
    )


def read_mapping(path: Path, required: tuple[str, ...], optional: tuple[str, ...],
                 extra_keys: tuple[str, ...] = ()) -> Mapping:
    text = path.read_text(encoding="utf-8")
    try:
        raw = tomllib.loads(text)
    except tomllib.TOMLDecodeError as exc:
        raise MappingError(f"{path}: {exc}") from exc

    unknown = set(raw) - BASE_KEYS - set(extra_keys)
    if unknown:
        raise MappingError(f"{path}: unknown settings {sorted(unknown)}")

    if ("sheet" in raw) == ("sheet_index" in raw):
        raise MappingError(
            f"{path}: give exactly one of sheet (its name) or sheet_index "
            "(1-based). Which sheet holds the data is not guessed."
        )

    columns_raw = raw.get("columns")
    if not isinstance(columns_raw, dict):
        raise MappingError(f"{path}: a [columns] table is required")

    known = set(required) | set(optional)
    unknown = set(columns_raw) - known
    if unknown:
        raise MappingError(
            f"{path}: [columns] has fields this importer does not know: "
            f"{sorted(unknown)}. Known fields: {sorted(known)}"
        )
    missing = [c for c in required if c not in columns_raw]
    if missing:
        raise MappingError(f"{path}: [columns] is missing {missing}")

    columns = {name: _column_index(name, value) for name, value in columns_raw.items()}
    doubled: dict[int, list[str]] = {}
    for name, index in columns.items():
        doubled.setdefault(index, []).append(name)
    clashes = {get_column_letter(i): n for i, n in doubled.items() if len(n) > 1}
    if clashes:
        raise MappingError(
            f"{path}: two fields point at the same column: {clashes}. "
            "That is almost always a mistake in the mapping."
        )

    first_data_row = raw.get("first_data_row")
    if not isinstance(first_data_row, int) or first_data_row < 1:
        raise MappingError(f"{path}: first_data_row is required, and is 1-based")
    last_data_row = raw.get("last_data_row")
    if last_data_row is not None and (
        not isinstance(last_data_row, int) or last_data_row < first_data_row
    ):
        raise MappingError(f"{path}: last_data_row must be at or after first_data_row")

    return Mapping(
        sheet=raw.get("sheet"),
        sheet_index=raw.get("sheet_index"),
        header_row=raw.get("header_row"),
        first_data_row=first_data_row,
        last_data_row=last_data_row,
        date_format=raw.get("date_format"),
        columns=columns,
        extra={key: raw[key] for key in extra_keys if key in raw},
        text=text,
        filename=str(path),
    )


def cell_text(value) -> str:
    """The cell as text, with nothing added and nothing removed.

    A number typed as a number arrives as an int or a float: `0090` in an
    unformatted Excel column is the number 90 and its leading zeros were lost
    in the spreadsheet, long before this.
    """
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else repr(value)
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat()
    return str(value)


def read_date(value, date_format: str | None, field_name: str, row: int,
              problems: list[RowProblem]) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = cell_text(value).strip()
    if not text:
        return None
    if not date_format:
        problems.append(RowProblem(
            row, field_name,
            f"{text!r} is text, not a date, and the mapping sets no date_format",
        ))
        return None
    try:
        return dt.datetime.strptime(text, date_format).date()
    except ValueError:
        problems.append(RowProblem(
            row, field_name, f"{text!r} does not fit date_format {date_format!r}"
        ))
        return None


def open_sheet(workbook, mapping: Mapping):
    if mapping.sheet is not None:
        if mapping.sheet not in workbook.sheetnames:
            raise MappingError(
                f"the workbook has no sheet named {mapping.sheet!r}. "
                f"It has: {workbook.sheetnames}"
            )
        return workbook[mapping.sheet]
    if not 1 <= mapping.sheet_index <= len(workbook.sheetnames):
        raise MappingError(
            f"sheet_index {mapping.sheet_index} is outside this workbook, "
            f"which has {len(workbook.sheetnames)} sheets"
        )
    return workbook[workbook.sheetnames[mapping.sheet_index - 1]]


def read_headers(sheet, mapping: Mapping) -> dict[str, str]:
    """Echoed in the report so a person can check the mapping. Never matched on."""
    if not mapping.header_row:
        return {}
    return {
        name: cell_text(sheet.cell(row=mapping.header_row, column=index).value).strip()
        for name, index in mapping.columns.items()
    }


def iter_rows(sheet, mapping: Mapping):
    """Yields (row_number, cells, blank). Blank rows are the caller's to skip."""
    last_row = mapping.last_data_row or sheet.max_row
    for row in range(mapping.first_data_row, last_row + 1):
        cells = {
            name: sheet.cell(row=row, column=index).value
            for name, index in mapping.columns.items()
        }
        blank = all(cell_text(v).strip() == "" for v in cells.values())
        yield row, cells, blank
