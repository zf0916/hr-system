#!/usr/bin/env python3
"""The gate for step 3: deliberate mistakes that must fail.

Every case runs in a transaction that is rolled back, so the gate can be run
against a live database without changing anything in it. Each case builds the
groups and schedules it needs, so nothing depends on what happens to be loaded.

    uv run python tools/schedule_gate.py

Exits non-zero if any deliberate mistake was accepted, or if a correct case was
refused.
"""

from __future__ import annotations

import datetime as dt
import shutil
import sys
import tempfile
from pathlib import Path

import psycopg
from sqlalchemy.exc import DatabaseError, IntegrityError

from app.calendar_import import adjust, describe_adjustments
from app.calendar_import import run_import as import_holidays
from app.db import Session
from sqlalchemy import func, select

from app.models import Employee, EmployeeGroup, GroupSchedule, Holiday
from app.schedule import schedule_for
from app.schedule import (
    attendance_day_for,
    day_context,
    effective_holiday,
    set_schedule,
)
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
HOLIDAYS = ROOT / "fixtures" / "holidays_provisional_2026.xlsx"
HOLIDAY_MAPPING = ROOT / "fixtures" / "holidays.mapping.toml"

DAY = {
    "start_time": dt.time(8, 0),
    "end_time": dt.time(17, 30),
    "end_next_day": False,
    "break_start": dt.time(12, 30),
    "break_end": dt.time(13, 15),
}
NIGHT = {
    "start_time": dt.time(19, 30),
    "end_time": dt.time(4, 30),
    "end_next_day": True,
}
COMMON = {"rest_weekdays": [7], "grace_minutes": 0, "provisional": True}


class Gate:
    def __init__(self) -> None:
        self.failures: list[str] = []
        self.checks = 0

    def check(self, ok: bool, what: str, detail: str = "") -> bool:
        self.checks += 1
        if ok:
            print(f"  ok      {what}")
        else:
            print(f"  FAIL    {what}" + (f" — {detail}" if detail else ""))
            self.failures.append(what)
        return ok

    def refused(self, what: str, fn) -> None:
        """The database, or the code, must refuse this."""
        with Session() as session:
            try:
                fn(session)
                session.flush()
            except (IntegrityError, DatabaseError, psycopg.Error, ValueError) as exc:
                first = str(exc).strip().splitlines()[0]
                self.check(True, what)
                print(f"            {first[:110]}")
            else:
                self.check(False, what, "it was accepted")
            finally:
                session.rollback()


def make_group(session, code: str) -> None:
    session.add(EmployeeGroup(code=code, label=code, note="gate"))
    session.flush()


def main() -> int:
    gate = Gate()

    print("\n-- overlapping schedules for one group")
    def overlapping(session):
        make_group(session, "GATE-DAY")
        session.add(GroupSchedule(
            group_code="GATE-DAY", effective_from=dt.date(2026, 1, 1),
            effective_to=dt.date(2026, 12, 31), **DAY, **COMMON))
        session.flush()
        # Straight into the table, not through set_schedule: the database is
        # what must refuse this, not whichever code path writes next.
        session.add(GroupSchedule(
            group_code="GATE-DAY", effective_from=dt.date(2026, 6, 1),
            effective_to=None, **DAY, **COMMON))
    gate.refused("two schedules covering the same day for one group", overlapping)

    print("\n-- a schedule that ends before it starts")
    def backwards_dates(session):
        make_group(session, "GATE-DAY")
        session.add(GroupSchedule(
            group_code="GATE-DAY", effective_from=dt.date(2026, 6, 1),
            effective_to=dt.date(2026, 1, 1), **DAY, **COMMON))
    gate.refused("effective_to before effective_from", backwards_dates)

    def night_without_saying_so(session):
        make_group(session, "GATE-NIGHT")
        session.add(GroupSchedule(
            group_code="GATE-NIGHT", effective_from=dt.date(2026, 1, 1),
            start_time=dt.time(19, 30), end_time=dt.time(4, 30),
            end_next_day=False, **COMMON))
    gate.refused(
        "a shift ending 04:30 after starting 19:30 without saying it crosses midnight",
        night_without_saying_so,
    )

    print("\n-- the same holiday date twice")
    def duplicate_holiday(session):
        for index in (1, 2):
            session.add(Holiday(
                holiday_date=dt.date(2026, 5, 1),
                name=f"Labour Day {index}", scope_code="federal",
                closes=True, provisional=True))
        session.flush()
    gate.refused("one date with two holiday rows", duplicate_holiday)

    with tempfile.TemporaryDirectory() as tmp:
        doubled = Path(tmp) / "doubled.xlsx"
        shutil.copy(HOLIDAYS, doubled)
        workbook = load_workbook(doubled)
        sheet = workbook["Holidays"]
        # Row 6 gets row 5's date: the same holiday typed twice, which is how
        # this arrives in a hand-filled sheet.
        sheet.cell(row=6, column=1).value = sheet.cell(row=5, column=1).value
        workbook.save(doubled)
        workbook.close()
        with Session() as session:
            result = import_holidays(
                session, doubled, HOLIDAY_MAPPING, year=2026, replace=True,
                allow_new=set(), provisional=True)
            messages = [str(p) for p in result.problems]
            session.rollback()
        gate.check(
            any("already on row" in m for m in messages),
            "the same date twice in the file",
            f"problems were {messages[:2]}",
        )

    with Session() as session:
        result = import_holidays(
            session, HOLIDAYS, HOLIDAY_MAPPING, year=2026, replace=True,
            allow_new=set(), provisional=True)
        problems = [str(p) for p in result.problems]
        session.rollback()
    gate.check(not problems, "the provisional holiday file loads",
               f"refused: {problems[:3]}")

    print("\n-- a past period must not move when today's schedule changes")
    with Session() as session:
        try:
            make_group(session, "GATE-DAY")
            set_schedule(session, "GATE-DAY", dt.date(2020, 1, 1), **DAY, **COMMON)
            past = dt.date(2026, 3, 10)
            before = day_context(session, "GATE-DAY", past)
            before_start, before_end = before.shift_start, before.shift_end
            before_row = before.schedule.id

            # A change from today: earlier start, later end, Saturday added as a
            # rest day. Everything a real change would touch.
            changed = dict(DAY)
            changed["start_time"] = dt.time(7, 0)
            changed["end_time"] = dt.time(19, 0)
            set_schedule(
                session, "GATE-DAY", dt.date(2026, 8, 18),
                **changed, rest_weekdays=[6, 7], grace_minutes=15, provisional=True)

            after = day_context(session, "GATE-DAY", past)
            gate.check(
                (after.shift_start, after.shift_end, after.schedule.id)
                == (before_start, before_end, before_row),
                "the past period renders with the row that was in force then",
                f"{before_start}–{before_end} became {after.shift_start}–"
                f"{after.shift_end}",
            )
            gate.check(
                after.is_rest_day is False,
                "the past period keeps the rest days it had, not today's",
            )
            today = day_context(session, "GATE-DAY", dt.date(2026, 8, 18))
            gate.check(
                today.shift_start.time() == dt.time(7, 0)
                and today.schedule.id != before_row,
                "today does use the new row",
                f"today is {today.shift_start}",
            )
        finally:
            session.rollback()

    print("\n-- a night-shift punch after midnight")
    with Session() as session:
        try:
            make_group(session, "GATE-NIGHT")
            set_schedule(session, "GATE-NIGHT", dt.date(2020, 1, 1), **NIGHT, **COMMON)
            tuesday_early = dt.datetime(2026, 8, 18, 4, 35, 0)
            landed = attendance_day_for(session, "GATE-NIGHT", tuesday_early)
            gate.check(
                landed.date == dt.date(2026, 8, 17),
                "04:35 on Tuesday belongs to Monday's attendance day",
                f"landed on {landed.date} ({landed.note})",
            )
            gate.check(landed.within_window, "and it is inside the shift window")

            start_of_shift = attendance_day_for(
                session, "GATE-NIGHT", dt.datetime(2026, 8, 17, 19, 25, 0))
            gate.check(
                start_of_shift.date == dt.date(2026, 8, 17),
                "19:25 on Monday belongs to Monday too",
                f"landed on {start_of_shift.date}",
            )
        finally:
            session.rollback()

    print("\n-- the same punch, with the crossing-midnight fact removed")
    with Session() as session:
        try:
            make_group(session, "GATE-NIGHT")
            # The database refuses 19:30–04:30 without end_next_day, so the way
            # to break this deliberately is a shift that genuinely does not
            # cross midnight. This is what step 6 would inherit if the row did
            # not carry the fact.
            set_schedule(
                session, "GATE-NIGHT", dt.date(2020, 1, 1),
                start_time=dt.time(19, 30), end_time=dt.time(23, 30),
                end_next_day=False, **COMMON)
            landed = attendance_day_for(
                session, "GATE-NIGHT", dt.datetime(2026, 8, 18, 4, 35, 0))
            gate.check(
                landed.date != dt.date(2026, 8, 17),
                "with a shift that does not cross midnight, 04:35 no longer "
                "lands on Monday — the check has teeth",
                "it still landed on Monday, so the test proves nothing",
            )
            print(f"            it landed on {landed.date} ({landed.note})")
        finally:
            session.rollback()

    print("\n-- re-loading one thing must not drop another")
    with Session() as session:
        try:
            make_group(session, "GATE-DAY")
            set_schedule(session, "GATE-DAY", dt.date(2020, 1, 1), **DAY, **COMMON)
            import_holidays(session, HOLIDAYS, HOLIDAY_MAPPING, year=2026,
                            replace=True, allow_new=set(), provisional=True)
            employees = session.scalar(
                select(func.count()).select_from(Employee))

            # Re-upload the calendar: schedules and employees untouched.
            import_holidays(session, HOLIDAYS, HOLIDAY_MAPPING, year=2026,
                            replace=True, allow_new=set(), provisional=True)
            gate.check(
                schedule_for(session, "GATE-DAY", dt.date(2026, 3, 10)) is not None,
                "re-uploading the calendar leaves the schedules alone",
            )
            gate.check(
                session.scalar(select(func.count()).select_from(Employee)) == employees,
                "re-uploading the calendar leaves the employee list alone",
            )

            # Change the rest day: the calendar is untouched.
            holidays_before = session.scalar(
                select(func.count()).select_from(Holiday))
            set_schedule(session, "GATE-DAY", dt.date(2026, 9, 1),
                         **DAY, rest_weekdays=[6, 7], grace_minutes=0,
                         provisional=True)
            gate.check(
                session.scalar(select(func.count()).select_from(Holiday))
                == holidays_before,
                "changing the rest day leaves the calendar alone",
            )
            gate.check(
                day_context(session, "GATE-DAY", dt.date(2026, 9, 5)).is_rest_day
                and not day_context(
                    session, "GATE-DAY", dt.date(2026, 3, 7)).is_rest_day,
                "the new rest day applies from its date, and not before it",
            )
        finally:
            session.rollback()

    print("\n-- not a gate: an adjustment, then a re-upload of the same year")
    with Session() as session:
        try:
            import_holidays(session, HOLIDAYS, HOLIDAY_MAPPING, year=2026,
                            replace=True, allow_new=set(), provisional=True)
            worked = dt.date(2026, 5, 1)
            adjust(session, date=worked, closes=False,
                   reason="line running to clear a shipment", made_by="Zi Fong")
            adjusted = effective_holiday(session, worked)
            gate.check(adjusted.closes is False and adjusted.adjusted,
                       "the adjustment changes the calendar")

            import_holidays(session, HOLIDAYS, HOLIDAY_MAPPING, year=2026,
                            replace=True, allow_new=set(), provisional=True)
            after = effective_holiday(session, worked)
            gate.check(
                after.closes is False and after.adjusted,
                "the adjustment survives a re-upload of the year",
                f"after re-upload the calendar says closes={after.closes}",
            )
            print("            re-upload report:")
            for date, verdict, detail in describe_adjustments(session, 2026):
                print(f"              {date} {verdict}: {detail}")
        finally:
            session.rollback()

    print(f"\n{gate.checks} checks")
    if gate.failures:
        print(f"{len(gate.failures)} FAILED:")
        for failure in gate.failures:
            print(f"  - {failure}")
        return 1
    print("clean")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
