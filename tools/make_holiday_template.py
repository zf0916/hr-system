#!/usr/bin/env python3
"""Builds the holiday spreadsheets in fixtures/.

Two files:

  holidays_template.xlsx           blank, with the columns HR fills in
  holidays_provisional_2026.xlsx   enough rows to exercise the importer

The 2026 list does not exist. The provisional file carries only holidays whose
dates are fixed by the calendar rather than declared each year, and every row
says PROVISIONAL in its notes. **Hari Raya, Chinese New Year, Deepavali, Wesak
and Thaipusam are deliberately absent** — their dates move, nobody here knows
them, and a plausible-looking wrong date is worse than a missing one.

    uv run python tools/make_holiday_template.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HEADERS = [
    ("A", "Date", 14),
    ("B", "Holiday Name", 42),
    ("C", "Scope", 12),
    ("D", "Company Closes?", 18),
    ("E", "Notes", 40),
]

INSTRUCTIONS = [
    "How to fill this in",
    "",
    "One row per date. Two holidays on the same date go on one row, with both",
    "names in Holiday Name — the calendar holds one row per date.",
    "",
    "Date             the date itself, typed as a date",
    "Holiday Name     what it is called on the notice",
    "Scope            federal, melaka, or company",
    "                 federal  gazetted for the whole country",
    "                 melaka   gazetted for Melaka state",
    "                 company  the factory closes, no government gazette",
    "Company Closes?  yes or no. NO is a real answer: a gazetted public holiday",
    "                 the factory works is exactly what this column is for.",
    "                 It decides whether the day is shaded on the sheet.",
    "Notes            anything the office should know. Optional.",
    "",
    "Rows may be in any order. Blank rows are ignored.",
    "The file is loaded a whole year at a time and a year can be re-loaded.",
    "A change to a single date after loading is made in the system, and survives",
    "a re-load.",
]

# Fixed-date holidays only. Everything that moves is left out on purpose.
PROVISIONAL_2026 = [
    ("2026-01-01", "New Year's Day", "federal", "yes", "PROVISIONAL — fixed date"),
    ("2026-04-15", "Declaration of Melaka as a Historical City", "melaka", "no",
     "PROVISIONAL — gazetted for Melaka; entered as a day the factory works, "
     "to exercise that case"),
    ("2026-05-01", "Labour Day", "federal", "yes", "PROVISIONAL — fixed date"),
    ("2026-08-31", "National Day", "federal", "yes", "PROVISIONAL — fixed date"),
    ("2026-09-16", "Malaysia Day", "federal", "yes", "PROVISIONAL — fixed date"),
    ("2026-12-25", "Christmas Day", "federal", "yes", "PROVISIONAL — fixed date"),
]

BANNER = PatternFill("solid", fgColor="FFF2CC")


def sheet_with_headers(workbook, title_line: str, banner_line: str):
    sheet = workbook.create_sheet("Holidays")
    sheet["A1"] = title_line
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A2"] = banner_line
    sheet["A2"].font = Font(bold=True, color="9C0006")
    sheet["A2"].fill = BANNER
    for column, header, width in HEADERS:
        cell = sheet[f"{column}4"]
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        sheet.column_dimensions[column].width = width
    return sheet


def instructions_sheet(workbook):
    sheet = workbook.active
    sheet.title = "How to fill this in"
    for index, line in enumerate(INSTRUCTIONS, start=1):
        sheet.cell(row=index, column=1, value=line)
    sheet["A1"].font = Font(bold=True, size=14)
    sheet.column_dimensions["A"].width = 80
    return sheet


def build_template() -> Workbook:
    workbook = Workbook()
    instructions_sheet(workbook)
    sheet_with_headers(
        workbook,
        "PUBLIC HOLIDAYS — <year>",
        "Blank template. Malaysia federal plus Melaka state. One row per date.",
    )
    return workbook


def build_provisional() -> Workbook:
    import datetime as dt

    workbook = Workbook()
    instructions_sheet(workbook)
    sheet = sheet_with_headers(
        workbook,
        "PUBLIC HOLIDAYS — 2026",
        "PROVISIONAL — NOT HR'S LIST. Fixed-date holidays only, for testing. "
        "Everything that moves (Hari Raya, CNY, Deepavali, Wesak, Thaipusam) is "
        "missing on purpose.",
    )
    for index, row in enumerate(PROVISIONAL_2026, start=5):
        date, name, scope, closes, note = row
        sheet.cell(row=index, column=1,
                   value=dt.datetime.strptime(date, "%Y-%m-%d").date())
        sheet.cell(row=index, column=1).number_format = "yyyy-mm-dd"
        sheet.cell(row=index, column=2, value=name)
        sheet.cell(row=index, column=3, value=scope)
        sheet.cell(row=index, column=4, value=closes)
        sheet.cell(row=index, column=5, value=note)
    return workbook


def main() -> int:
    fixtures = Path(__file__).resolve().parent.parent / "fixtures"
    fixtures.mkdir(parents=True, exist_ok=True)
    build_template().save(fixtures / "holidays_template.xlsx")
    build_provisional().save(fixtures / "holidays_provisional_2026.xlsx")
    print(f"wrote {fixtures / 'holidays_template.xlsx'}")
    print(f"wrote {fixtures / 'holidays_provisional_2026.xlsx'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
