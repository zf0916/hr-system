#!/usr/bin/env python3
"""Write an employee list of a stated size, shaped like HR's, for a rehearsal.

**Headcount is unread** (SPEC §9 A39). The committed demonstration list has 58
rows because that is how many punches were captured to attach to — it says
nothing about how many people work at the factory, and its numbers are invented
apart from three. The one real datum is that HR's own paper prints `1601`, which
bounds the numbers *issued* over the years, not the people employed now.

So this tool takes the size as an argument and says so on the sheet's first
line. It exists to answer questions the 58-row list cannot: how many printed
pages, how long a render takes, whether the sheet screen is usable when the rows
run past the bottom of it.

**Not HR's list, and not a second fixture to be believed.** The output belongs
in `import/`, which is not committed, and it is loaded onto a throwaway database
— never the working one, which holds leave records, gate passes and corrections
that a `--replace` is refused in front of (SPEC §6).

Numbers are spread across 1–1600 with gaps, because that is what a numbering
that has reached 1601 looks like after people have left. **PINs carry no leading
zero** — the device refuses one (SPEC §10) — so an employee numbered `0090` gets
PIN `90`, which is what the dated device-user mapping is for.

    uv run python tools/make_size_fixture.py --count 300 \
        --out import/employees_size300.xlsx
"""

from __future__ import annotations

import argparse
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

# The same column letters as the committed sample, so the same mapping file
# reads this. Nothing here is matched on header text (SPEC §2).
HEADERS = [
    (1, "No."), (2, "Emp No."), (4, "Employee Name"), (5, "Section"),
    (6, "Role"), (7, "Shift Group"), (9, "Joined"), (10, "Left"),
    (11, "Device ID"),
]

SECTIONS = ["PACK ASSY", "QC", "MAINT", "PROJECT DOOR", "WAREHOUSE"]
ROLES = ["Production Assistant", "QA/QC", "Charge Hand", "Assistant Supervisor",
         "HOD/Supervisor", "Management/Office"]
GROUPS = ["DAY-PROD", "NIGHT-PROD", "OFFICE"]

# Distinct names, because a sheet read by eye with two people called the same
# thing on it is a mistake waiting to happen. 60 x 20 = 1200 pairs.
FIRST = [
    "Ah Meng", "Siti", "Ravi", "Wei Sheng", "Nurul", "Chee Keong", "Mei Ling",
    "Aung", "Chandran", "Zubaidah", "Kai Xin", "Hafiz", "Su Lin", "Arun",
    "Boon Hui", "Farah", "Jia Hui", "Kumar", "Lina", "Tze Yang", "Poh Choo",
    "Zainal", "Vimala", "Beng Huat", "Rosnah", "Sanjay", "Yoke Lan", "Idris",
    "Kalai", "Cheng Hoe", "Norhayati", "Prakash", "Swee Lian", "Faizal",
    "Devi", "Kok Wai", "Anisah", "Ganesan", "Peik Ying", "Rahim",
    "Azman", "Bee Choo", "Cheng Wai", "Dhanraj", "Elango", "Fadzil",
    "Guan Eng", "Hasnah", "Indra", "Jamal", "Kok Leong", "Latifah",
    "Manoharan", "Noraini", "Ong Kim", "Puspa", "Ramlah", "Selvam",
    "Tuan Mat", "Wan Ling",
]
LAST = [
    "Tan", "binti Rahman", "a/l Subramaniam", "Lim", "Wong", "Ko Min",
    "a/l Muthu", "binti Osman", "Chong", "Ismail", "Lee", "Ng", "a/p Raju",
    "bin Salleh", "Cheah", "Yap", "Loh", "a/l Nathan", "binti Yusof", "Teoh",
]

# The highest number HR's own paper is known to carry, plus room. Numbers are
# drawn from below this with gaps (SPEC §2).
NUMBER_CEILING = 1600


def name_for(index: int) -> str:
    return f"{FIRST[index % len(FIRST)]} {LAST[(index // len(FIRST)) % len(LAST)]}"


def rows(count: int, seed: int) -> list[tuple]:
    if count > len(FIRST) * len(LAST):
        raise SystemExit(
            f"{count} is more than the {len(FIRST) * len(LAST)} distinct names "
            "this generator can make. Add names before asking for more people")
    if count > NUMBER_CEILING:
        raise SystemExit(f"{count} is more than the {NUMBER_CEILING} numbers "
                         "below the ceiling HR's paper shows")

    rng = random.Random(seed)
    numbers = sorted(rng.sample(range(1, NUMBER_CEILING + 1), count))

    out: list[tuple] = []
    for index, number in enumerate(numbers):
        if index % 7 == 0:
            group = "NIGHT-PROD"
        elif index % 11 == 0:
            group = "OFFICE"
        else:
            group = "DAY-PROD"
        out.append((
            f"{number:04d}",                       # stored padded, as §2 has it
            name_for(index),
            SECTIONS[index % len(SECTIONS)],
            ROLES[index % len(ROLES)],
            group,
            "05/01/2020",
            "",
            str(number),                           # no leading zero: SPEC §10
        ))
    return out


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--count", type=int, required=True,
                        help="how many employees. Headcount is unread (A39); "
                             "this is a rehearsal size somebody chose")
    parser.add_argument("--out", required=True)
    parser.add_argument("--seed", type=int, default=20260822)
    args = parser.parse_args()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Staff List"

    sheet["A1"] = (f"REHEARSAL LIST — {args.count} invented employees. "
                   "Not HR's list, and headcount is unread (SPEC §9 A39)")
    sheet["A1"].font = Font(bold=True)
    sheet["A2"] = "Generated by tools/make_size_fixture.py"

    header_row = 4
    for column, text in HEADERS:
        sheet.cell(header_row, column, text).font = Font(bold=True)

    data = rows(args.count, args.seed)
    names = [row[1] for row in data]
    duplicates = sorted({name for name in names if names.count(name) > 1})
    if duplicates:
        raise SystemExit(f"{len(duplicates)} repeated name(s): {duplicates[:5]}")

    for index, row in enumerate(data):
        excel_row = header_row + 1 + index
        number, name, section, role, group, joined, left, pin = row
        sheet.cell(excel_row, 1, index + 1)
        sheet.cell(excel_row, 2, number)
        sheet.cell(excel_row, 4, name)
        sheet.cell(excel_row, 5, section)
        sheet.cell(excel_row, 6, role)
        sheet.cell(excel_row, 7, group)
        sheet.cell(excel_row, 9, joined)
        sheet.cell(excel_row, 10, left)
        sheet.cell(excel_row, 11, pin)

    footer = header_row + 1 + len(data)
    sheet.cell(footer, 2, "TOTAL")
    sheet.cell(footer, 4, f"{len(data)} employees")

    for letter, width in (("A", 6), ("B", 10), ("D", 26), ("E", 16), ("F", 22),
                          ("G", 14), ("I", 12), ("J", 12), ("K", 12)):
        sheet.column_dimensions[letter].width = width

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out)

    # The mapping is written beside the file rather than kept by hand, because
    # `last_data_row` moves with the count and a mapping one row out either
    # imports the TOTAL footer or drops the last employee.
    mapping = out.with_suffix("").with_suffix(".mapping.toml")
    mapping.write_text(
        f"# Written by tools/make_size_fixture.py beside {out.name}.\n"
        "# A rehearsal list, not HR's. Column letters are explicit and header\n"
        "# text is never matched on (SPEC §2).\n\n"
        'sheet = "Staff List"\n'
        f"header_row = {header_row}\n"
        f"first_data_row = {header_row + 1}\n"
        f"last_data_row = {footer - 1}        # {footer} is the TOTAL footer\n\n"
        'date_format = "%d/%m/%Y"\n\n'
        "[columns]\n"
        'employee_number = "B"\n'
        'name            = "D"\n'
        'section         = "E"\n'
        'role            = "F"\n'
        'group           = "G"\n'
        'active_from     = "I"\n'
        'left_on         = "J"\n'
        'device_pin      = "K"\n'
    )

    print(f"wrote {out}: {len(data)} employees, numbers "
          f"{data[0][0]}–{data[-1][0]}, header row {header_row}, "
          f"data rows {header_row + 1}-{footer - 1}, footer row {footer}")
    print(f"  and {mapping}")
    print("  every PIN is the number without its padding — the device refuses "
          "a leading zero (SPEC §10)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
