#!/usr/bin/env python3
"""Read an exported sheet back and compare it to the render that produced it.

**This is a check on the writer, not an ingest path.** Nothing in `app/` reads a
sheet file — the file goes one way, and a returned sheet would be a correction
in a cell instead of a row (SPEC §7, §13). This lives in `tools/` for that
reason: it exists to prove that the Excel says what the screen says, and it
never writes anything anywhere.

    uv run python tools/sheet_readback.py --month 2026-08 --file sheet.xlsx
"""

from __future__ import annotations

import argparse
import datetime as dt
import sys

from openpyxl import load_workbook

from app.db import Session
from app.sheet import excel_layout, page_layout, period_for, render, to_text


def read_sheet_file(path) -> dict:
    """What the file actually contains: the title, the period line, the day
    columns, the note state, and every cell keyed by (employee number, day)."""
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook.active
    layout = excel_layout()

    days: list[int] = []
    column = layout["first_day_column"]
    while True:
        value = worksheet.cell(layout["header_row"], column).value
        if value is None:
            break
        days.append(int(value))
        column += 1

    cells: dict[tuple[str, int], str] = {}
    numbers: list[str] = []
    row = layout["first_data_row"]
    while True:
        number = worksheet.cell(row, layout["number_column"]).value
        if number in (None, "", "Legend"):
            break
        number = str(number)
        numbers.append(number)
        for index, day in enumerate(days):
            value = worksheet.cell(row, layout["first_day_column"] + index).value
            cells[(number, day)] = "" if value is None else str(value)
        row += 1

    note_cell = worksheet["A4"].value
    return {
        # How the file prints. The Excel file is the record HR files (SPEC §7),
        # so a missing page setup is a defect in the artefact even when every
        # cell is right.
        "orientation": worksheet.page_setup.orientation,
        "fit_to_width": worksheet.page_setup.fitToWidth,
        "fit_to_height": worksheet.page_setup.fitToHeight,
        "fit_to_page": bool(getattr(worksheet.sheet_properties.pageSetUpPr,
                                    "fitToPage", None)),
        "print_title_rows": worksheet.print_titles,
        "row_breaks": sorted(brk.id for brk in worksheet.row_breaks.brk),
        "title": worksheet["A1"].value,
        "period": worksheet["A2"].value,
        "headcount_line": worksheet["A3"].value,
        "note": "" if note_cell is None else str(note_cell),
        "note_marker": worksheet["B4"].value,
        "days": days,
        "numbers": numbers,
        "cells": cells,
    }


def compare(sheet, file_contents: dict) -> list[str]:
    """Every difference between the render and the file. Empty means they agree
    about every day for every employee."""
    problems: list[str] = []
    if file_contents["title"] != sheet.title:
        problems.append(
            f"title: file {file_contents['title']!r} vs render {sheet.title!r}")
    expected_period = f"{sheet.period_start} to {sheet.period_end}"
    if file_contents["period"] != expected_period:
        problems.append(
            f"period: file {file_contents['period']!r} vs {expected_period!r}")
    if file_contents["days"] != [c.day for c in sheet.columns]:
        problems.append(
            f"day columns: file {file_contents['days']} vs "
            f"{[c.day for c in sheet.columns]}")
    if file_contents["numbers"] != [r.employee_number for r in sheet.rows]:
        problems.append(
            f"employee rows: file {file_contents['numbers']} vs "
            f"{[r.employee_number for r in sheet.rows]}")
    if sheet.note_is_unread:
        if file_contents["note"].strip():
            problems.append(
                f"the unread note came out with content: {file_contents['note']!r}")
        if not (file_contents["note_marker"] or "").strip():
            problems.append("the unread note is not marked as unread in the file")

    page = page_layout(sheet)
    if str(file_contents["orientation"]) != page["orientation"]:
        problems.append(
            f"orientation: file {file_contents['orientation']!r} vs "
            f"{page['orientation']!r} — the file is printed, not read on screen")
    if file_contents["fit_to_width"] != page["fit_to_width"]:
        problems.append(
            f"fit to width: file {file_contents['fit_to_width']!r} vs "
            f"{page['fit_to_width']!r} — 31 day columns spilling onto a second "
            "page wide is unreadable")
    if file_contents["fit_to_height"] != page["fit_to_height"]:
        problems.append(
            f"fit to height: file {file_contents['fit_to_height']!r} vs "
            f"{page['fit_to_height']!r}")
    if not file_contents["fit_to_page"]:
        problems.append("fit-to-page is not switched on, so the fit settings "
                        "are ignored by Excel")
    # Excel writes the titles qualified and absolute: "'Attendance'!$6:$7".
    titles = (file_contents["print_title_rows"] or "").split("!")[-1].replace("$", "")
    if titles != page["print_title_rows"]:
        problems.append(
            f"repeating title rows: file {file_contents['print_title_rows']!r} "
            f"vs rows {page['print_title_rows']} — without them, every page but "
            "the first is ticks with no day numbers above them")
    if file_contents["row_breaks"] != page["row_breaks"]:
        problems.append(
            f"page breaks: file {file_contents['row_breaks']} vs "
            f"{page['row_breaks']} — {page['rows_per_page']} rows to a page "
            "(SPEC §9 A39), and the legend on its own page at the end")

    for row in sheet.rows:
        for column in sheet.columns:
            rendered = sheet.cell(row.employee_id, column.date).text
            in_file = file_contents["cells"].get(
                (row.employee_number, column.day))
            if in_file is None:
                problems.append(
                    f"{row.employee_number} {column.date}: missing from the file")
            elif in_file != rendered:
                problems.append(
                    f"{row.employee_number} {column.date}: file {in_file!r} vs "
                    f"screen {rendered!r}")
    return problems


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--file", required=True)
    parser.add_argument("--month")
    parser.add_argument("--from", dest="start")
    parser.add_argument("--to", dest="end")
    parser.add_argument("--section")
    parser.add_argument("--print", action="store_true",
                        help="print what the file contains, cell by cell")
    args = parser.parse_args()

    with Session() as session:
        if args.month:
            start, end = period_for(session, args.month)
        else:
            start = dt.datetime.strptime(args.start, "%Y-%m-%d").date()
            end = dt.datetime.strptime(args.end, "%Y-%m-%d").date()
        sheet = render(session, start, end, section_code=args.section)

    contents = read_sheet_file(args.file)
    problems = compare(sheet, contents)

    print(f"file:   {args.file}")
    print(f"title:  {contents['title']!r}")
    print(f"period: {contents['period']!r}")
    print(f"note:   {contents['note']!r}   marker: {contents['note_marker']!r}")
    print(f"days:   {len(contents['days'])}   employees: "
          f"{len(contents['numbers'])}")
    print(f"print:  {contents['orientation']}, fit to width "
          f"{contents['fit_to_width']}, titles {contents['print_title_rows']!r}, "
          f"{len(contents['row_breaks'])} page break(s) at "
          f"{contents['row_breaks']}")
    filled = {k: v for k, v in contents["cells"].items() if v}
    print(f"cells with something in them: {len(filled)}")
    for (number, day), value in sorted(filled.items()):
        print(f"  {number}  day {day:>2}  {value!r}")
    if args.print:
        print("\n-- the screen, for comparison")
        print(to_text(sheet))

    if problems:
        print(f"\n{len(problems)} DIFFERENCES between the file and the render:")
        for problem in problems:
            print(f"  - {problem}")
        return 1
    print("\nthe file and the render agree about every day for every employee")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
