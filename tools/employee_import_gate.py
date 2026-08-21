#!/usr/bin/env python3
"""The gate for the employee importer: deliberate mistakes that must fail.

A checker that only ever sees good input has not been tested. Each case below
breaks the fixture or the mapping in a way a real load could plausibly be
broken, and the case passes only if the import refuses it *and* says something
a person could act on.

Nothing is committed. Every case runs inside a transaction that is rolled back,
so the gate can be run against a database with the list already loaded.

**It runs against a throwaway database, and has to.** Most cases need a
`--replace`, which is refused while a leave record, a gate pass or a correction
exists — and a correction can no longer be cleared out of the way, because the
database refuses to delete one (SPEC §3, §13). So the run happens on a database
created for it and dropped at the end: `tools/throwaway.py` starts the same
image against a scratch database on the compose network, and this file re-runs
itself inside it. `--inside` is that second run.

    uv run python tools/employee_import_gate.py
    uv run python tools/employee_import_gate.py --inside   # in the container

Exits non-zero if any deliberate mistake was accepted, or if the clean fixture
was refused.
"""

from __future__ import annotations

import shutil
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from sqlalchemy import text

from app.db import Session
from app.employee_import import MappingError, run_import

ROOT = Path(__file__).resolve().parent.parent
FIXTURE = ROOT / "fixtures" / "employees_sample.xlsx"

BASE_MAPPING = {
    "sheet": '"Staff List"',
    "header_row": "4",
    "first_data_row": "5",
    "last_data_row": "12",
    "date_format": '"%d/%m/%Y"',
}
BASE_COLUMNS = {
    "employee_number": "B",
    "name": "D",
    "section": "E",
    "role": "F",
    "group": "G",
    "active_from": "I",
    "left_on": "J",
    "device_pin": "K",
}


def write_mapping(path: Path, top: dict | None = None,
                  columns: dict | None = None) -> Path:
    settings = dict(BASE_MAPPING)
    settings.update(top or {})
    cols = dict(BASE_COLUMNS)
    cols.update(columns or {})
    lines = [f"{key} = {value}" for key, value in settings.items() if value is not None]
    lines.append("")
    lines.append("[columns]")
    lines += [f'{key} = "{value}"' for key, value in cols.items() if value is not None]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def write_workbook(path: Path, edits: list[tuple[int, int, object]]) -> Path:
    shutil.copy(FIXTURE, path)
    if not edits:
        return path
    workbook = load_workbook(path)
    sheet = workbook["Staff List"]
    for row, column, value in edits:
        # openpyxl's cell(value=...) ignores None, so blanking is an empty string.
        sheet.cell(row=row, column=column).value = value
    workbook.save(path)
    workbook.close()
    return path


class Gate:
    def __init__(self) -> None:
        self.failures: list[str] = []
        self.checks = 0
        self.last_staged: list = []
        self.last_headers: dict = {}

    def check(self, ok: bool, what: str, detail: str = "") -> None:
        self.checks += 1
        if ok:
            print(f"  ok      {what}")
        else:
            print(f"  FAIL    {what}" + (f" — {detail}" if detail else ""))
            self.failures.append(what)

    def attempt(self, workdir: Path, label: str, *, edits=None, top=None,
                columns=None, flags=(), allow_new=("group",),
                preload=False, keep_hr_entry=False) -> tuple[list[str], dict]:
        """Run one import. Returns (messages, written) and never commits."""
        source = write_workbook(workdir / f"{label}.xlsx", edits or [])
        mapping = write_mapping(workdir / f"{label}.toml", top, columns)
        with Session() as session:
            try:
                if keep_hr_entry == "guard":
                    # A guard entry is not a typed form, and is refused for a
                    # stronger reason: §13 forbids deleting a correction at
                    # all. Before this was checked, `--replace` did not refuse
                    # — it reached the DELETE and died on a foreign key.
                    session.execute(text(
                        "INSERT INTO manual_punch (employee_id, path, "
                        "attendance_day, reason_code, made_by) SELECT id, "
                        "'guard', '2026-08-21', 'biometric_failed', 'gate' "
                        "FROM employee LIMIT 1"))
                elif keep_hr_entry:
                    # The case that proves a --replace is refused while HR
                    # entry exists needs one leave record to exist. It makes
                    # its own, here, inside the transaction that is rolled
                    # back — **never against whatever HR has typed.** A leave
                    # record is a form somebody signed; it is not rebuildable
                    # from anything, and a gate that clears the table to make
                    # room for itself destroys exactly the data this case is
                    # about (SPEC §5, §6).
                    session.execute(text(
                        "INSERT INTO leave_record (employee_id, "
                        "leave_type_code, sheet_code, period_from, period_to, "
                        "days, entered_by) SELECT id, 'ANNUAL', 'AL', "
                        "'2026-08-24', '2026-08-24', 1, 'gate' FROM employee "
                        "LIMIT 1"))
                # Every other case tests the list itself, and needs a
                # database with nothing a person recorded in the way. **It gets
                # one by being a throwaway**, not by clearing tables: a leave
                # record is a form somebody signed, and a correction cannot be
                # deleted at all — the database refuses it (SPEC §3, §5, §6,
                # §13). The two cases above make their own row and roll it
                # back, which is the only way rows get here.
                session.flush()
                if preload:
                    run_import(session, FIXTURE,
                               write_mapping(workdir / f"{label}-pre.toml"),
                               replace=True, allow_new={"group"},
                               accept_odd_numbers=False)
                result = run_import(
                    session,
                    source,
                    mapping,
                    replace="no-replace" not in flags and not preload,
                    allow_new=set(allow_new),
                    accept_odd_numbers="accept-odd-numbers" in flags,
                    accept_leading_zero_pins="accept-leading-zero-pins" in flags,
                )
            except MappingError as exc:
                session.rollback()
                return [str(exc)], {}
            written = dict(result.written)
            self.last_staged = list(result.staged)
            self.last_headers = dict(result.headers)
            messages = [f"{p.field}: {p.message}" for p in result.problems]
            session.rollback()
            return messages, written if not result.problems else {}

    def must_fail(self, workdir: Path, label: str, expected: str, **kwargs) -> None:
        messages, written = self.attempt(workdir, label, **kwargs)
        if not messages:
            self.check(False, label, f"accepted it and wrote {written}")
            return
        hit = any(expected.lower() in m.lower() for m in messages)
        self.check(hit, label, f"failed, but for another reason: {messages[:2]}")

    def must_pass(self, workdir: Path, label: str, expect_employees: int,
                  **kwargs) -> dict:
        messages, written = self.attempt(workdir, label, **kwargs)
        if messages:
            self.check(False, label, f"refused it: {messages[:3]}")
            return {}
        self.check(
            written.get("employee") == expect_employees,
            label,
            f"wrote {written.get('employee')} employees, expected {expect_employees}",
        )
        return written


def main() -> int:
    gate = Gate()
    if not FIXTURE.exists():
        print(f"no fixture at {FIXTURE} — run tools/make_employee_fixture.py",
              file=sys.stderr)
        return 2

    with tempfile.TemporaryDirectory() as tmp:
        work = Path(tmp)

        print("\n-- the fixture, mapped correctly")
        gate.must_pass(work, "clean fixture loads", 8)

        print("\n-- a mapping pointing at the wrong column")
        # This used to be caught by the shape: a running number 1..8 is not
        # four digits. It is not caught that way any more, because HR's paper
        # writes real numbers short and `090` has to load (SPEC §2). **What
        # shows it is the header echo**, which is why the importer reads the
        # header row at all — it never matches on it, it prints it back.
        written = gate.must_pass(
            work, "employee_number pointed at the running-number column loads",
            8, columns={"employee_number": "A"})
        if written:
            header = gate.last_headers.get("employee_number")
            gate.check(header not in (None, "Emp No."),
                       "and the report says which column it read: "
                       f"employee_number <- {header!r}, not the number column",
                       f"headers {gate.last_headers}")
            numbers = [row.employee_number for row in gate.last_staged]
            gate.check(numbers == [str(i) for i in range(1, 9)],
                       f"the numbers it loaded are the running numbers: {numbers}")
            gate.check(all(row.number_from_numeric_cell
                           for row in gate.last_staged),
                       "and every one came from a numeric cell, which the "
                       "report also says")
        gate.must_fail(
            work, "employee_number pointed at the name column", "expected shape",
            columns={"employee_number": "D", "name": "B"},
        )
        gate.must_fail(
            work, "section and name swapped", "is not a known section",
            columns={"section": "D", "name": "E"},
        )
        gate.must_fail(
            work, "active_from and name swapped", "does not fit date_format",
            columns={"active_from": "D", "name": "I"},
        )
        gate.must_fail(
            work, "name pointed at the remarks column", "name: is empty",
            columns={"name": "C"},
        )
        gate.must_fail(
            work, "two fields pointed at one column",
            "point at the same column",
            columns={"name": "E"},
        )

        print("\n-- the data itself")
        gate.must_fail(
            work, "a duplicate employee number", "is already used on row",
            edits=[(7, 2, "0090")],
        )
        gate.must_fail(
            work, "two numbers that pad to the same key", "the same as row",
            edits=[(7, 2, "90")], flags=("accept-odd-numbers",),
        )
        gate.must_fail(
            work, "a left date before the active date", "is before active_from",
            edits=[(8, 10, "01/01/2019")],
        )
        gate.must_fail(
            work, "a number that cannot be keyed to four", "expected shape",
            edits=[(5, 2, "16011")],
        )
        gate.must_fail(
            work, "a number that is not digits", "expected shape",
            edits=[(5, 2, "A090")],
        )
        gate.must_fail(
            work, "one PIN on two employees", "cannot belong to two employees",
            edits=[(6, 11, "90")],
        )
        gate.must_fail(
            work, "an empty name", "name: is empty",
            edits=[(9, 4, "")],
        )
        gate.must_fail(
            work, "a date that does not fit the stated format",
            "does not fit date_format",
            edits=[(9, 9, "March 2019")],
        )

        print("\n-- the mapping's own shape")
        gate.must_fail(
            work, "a sheet name that is not in the workbook", "has no sheet named",
            top={"sheet": '"Staff Listing"'},
        )
        gate.must_fail(
            work, "the data starting on the header row", "expected shape",
            top={"first_data_row": "4"},
        )
        gate.must_fail(
            work, "no last_data_row, so the footer is read", "is empty",
            top={"last_data_row": None},
        )
        gate.must_fail(
            work, "a required field left out of the mapping", "is missing",
            columns={"group": None},
        )
        gate.must_fail(
            work, "a column given as a header name instead of a letter",
            "is not a column letter or number",
            columns={"name": "Employee Name"},
        )
        gate.must_fail(
            work, "text dates with no date_format stated",
            "no date_format",
            top={"date_format": None},
        )

        print("\n-- vocabulary and reloading")
        gate.must_fail(
            work, "a new group without --allow-new group", "is not a known group",
            edits=[(5, 7, "GROUP-NOBODY-HAS-SEEN")], allow_new=(),
        )
        gate.must_fail(
            work, "--allow-new group does not also allow new sections",
            "is not a known section",
            edits=[(5, 5, "SECTION-NOBODY-HAS-SEEN")],
        )
        gate.must_fail(
            work, "loading a second time without --replace", "is already loaded",
            flags=("no-replace",), preload=True,
        )

        print("\n-- a number written short is ordinary, not odd (SPEC §2)")
        # HR's own paper prints `090` and `1601` on one page. A three-digit
        # number is the same number written short, so it loads with no flag and
        # keys to four.
        written = gate.must_pass(
            work, "a three-digit number loads with no flag", 8,
            edits=[(5, 2, "090")],
        )
        if written:
            staged = {row.employee_number: row for row in gate.last_staged}
            gate.check("090" in staged and staged["090"].number_key == "0090",
                       "stored verbatim as '090', keyed as '0090'",
                       f"got {[(n, r.number_key) for n, r in staged.items()][:3]}")
            gate.check(not staged["090"].odd_number,
                       "and it is not flagged odd — nothing had to be accepted")

        gate.must_pass(work, "so does a two-digit one", 8, edits=[(5, 2, "90")])

        print("\n-- what an odd number does when it is accepted deliberately")
        written = gate.must_pass(
            work, "--accept-odd-numbers loads a five-digit number", 8,
            edits=[(5, 2, "16011")],
            flags=("accept-odd-numbers",),
        )
        if written:
            print("           stored verbatim, and its key is the number itself "
                  "— five digits cannot pad to four")

    print("\n-- HR entry is not cleared with the employee list (SPEC §5)")
    with tempfile.TemporaryDirectory() as directory:
        work = Path(directory)
        # Both cases make and discard their own leave record inside the
        # transaction. **Nothing here commits**, so HR's own entry — which
        # exists on paper and nowhere else — is not touched.
        gate.must_fail(
            work, "--replace is refused while HR entry exists",
            expected="cannot be rebuilt from anything",
            keep_hr_entry=True)
        # The same import, with the leave record out of the way: it is the
        # HR entry that refuses it, not anything about the list.
        gate.must_fail(
            work, "and refused while a guard's correction exists, by name",
            expected="manual_punch",
            keep_hr_entry="guard")
        # The same import, with the leave record out of the way: it is the
        # HR entry that refuses it, not anything about the list.
        gate.must_pass(
            work, "and the same --replace runs once the HR entry is gone", 8,
            keep_hr_entry=False)

    print("\n-- a device PIN the hardware cannot hold (SPEC §2, §10)")
    with tempfile.TemporaryDirectory() as directory:
        work = Path(directory)
        gate.must_fail(
            work, "a PIN with a leading zero is refused",
            expected="leading zero",
            edits=[(6, 11, "0142")],
        )
        written = gate.must_pass(
            work, "and loads when it is asked for deliberately", 8,
            edits=[(6, 11, "0142")],
            flags=("accept-leading-zero-pins",),
        )
        if written:
            staged = {row.employee_number: row for row in gate.last_staged}
            gate.check(staged["0142"].device_pin == "0142",
                       "stored exactly as the list gave it")
            gate.check("no punch" in staged["0142"].pin_note,
                       "and the row says no punch can ever carry it",
                       f"note {staged['0142'].pin_note!r}")

        # A single zero is a PIN the device can hold — 0 is not a leading zero.
        gate.must_pass(work, "a PIN of '0' is not a leading-zero PIN", 8,
                       edits=[(6, 11, "0")])

    print("\n-- every employee's PIN outcome is reported, not just a count")
    with tempfile.TemporaryDirectory() as directory:
        work = Path(directory)
        written = gate.must_pass(work, "the fixture loads", 8)
        staged = {row.employee_number: row for row in gate.last_staged}
        gate.check(len(staged) == 8, "eight employees staged",
                   f"got {len(staged)}")
        mapped = [n for n, r in staged.items() if r.device_pin]
        blank = [n for n, r in staged.items() if not r.device_pin]
        gate.check(len(mapped) == 6 and len(blank) == 2,
                   f"six mapped, two without a PIN: {sorted(blank)}",
                   f"mapped {sorted(mapped)}, blank {sorted(blank)}")
        gate.check(all("empty" in staged[n].pin_note for n in blank),
                   "and each says why, in words",
                   f"{[staged[n].pin_note for n in blank]}")

    print(f"\n{gate.checks} checks")
    if gate.failures:
        print(f"{len(gate.failures)} FAILED:")
        for failure in gate.failures:
            print(f"  - {failure}")
        return 1
    print("clean")
    return 0


def outside() -> int:
    """Start a throwaway database, and run this gate inside it.

    The same image, on the same compose network, against a database created for
    this run and dropped when it ends. Its output is passed straight through,
    so what a reader sees is what the gate printed.
    """
    import subprocess

    sys.path.insert(0, str(ROOT))
    from tools.throwaway import CONTAINER, Throwaway

    with Throwaway() as scratch:
        scratch.run("hr", "seed", "--add-missing")
        result = subprocess.run(
            ["docker", "exec", CONTAINER, "python",
             "tools/employee_import_gate.py", "--inside"],
            text=True)
        return result.returncode


if __name__ == "__main__":
    raise SystemExit(main() if "--inside" in sys.argv else outside())
