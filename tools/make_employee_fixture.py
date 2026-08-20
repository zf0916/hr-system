#!/usr/bin/env python3
"""Builds fixtures/employees_sample.xlsx.

The real list has not arrived and nobody has seen its columns, so this fixture
is deliberately awkward in the ways a real one is: three sheets, a title block
above the header row, a running-number column that looks like an employee
number, a stray Remarks column, a gap column, dates written as text, and PINs
that do not all match the employee number.

The headers are our own invention. Nothing in the importer reads them.

    uv run python tools/make_employee_fixture.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

HEADERS = [
    (1, "S/N"),
    (2, "No."),
    (3, "Remarks"),
    (4, "Employee Name"),
    (5, "Dept"),
    (6, "Category"),
    (7, "Shift Group"),
    (9, "Date Joined"),
    (10, "Date Left"),
    (11, "Device ID"),
]

# number, remarks, name, dept, category, group, joined, left, device id
ROWS = [
    ("0090", "", "Lim Wei Sheng", "PACK ASSY", "Production Assistant",
     "DAY-PROD", "12/03/2018", "", "90"),
    # PIN 142, not 0142: the device refuses a leading zero in a user ID
    # (SPEC §2, §10), and a fixture carrying one would ship a mapping row that
    # can never match a punch. The refusal itself is proven in the gate.
    ("0142", "", "Nurul Aisyah binti Rahman", "QC", "QA/QC",
     "DAY-PROD", "05/07/2019", "", "142"),
    ("0657", "", "Tan Chee Keong", "MAINT", "Charge Hand",
     "DAY-PROD", "21/11/2015", "", "657"),
    ("0881", "resigned", "Ravi a/l Subramaniam", "WAREHOUSE", "Production Assistant",
     "NIGHT-PROD", "02/01/2020", "30/06/2026", "881"),
    ("1042", "", "Siti Zubaidah binti Osman", "PROJECT DOOR", "Assistant Supervisor",
     "DAY-PROD", "17/09/2021", "", ""),
    ("1288", "", "Wong Mei Ling", "QC", "Management/Office",
     "OFFICE", "04/04/2016", "", "1288"),
    ("1627", "", "Aung Ko Min", "PACK ASSY", "Production Assistant",
     "NIGHT-PROD", "13/02/2023", "", "1627"),
    ("1903", "night shift", "Chandran a/l Muthu", "MAINT", "HOD/Supervisor",
     "NIGHT-PROD", "08/08/2024", "", ""),
]


def build() -> Workbook:
    workbook = Workbook()

    cover = workbook.active
    cover.title = "Cover"
    cover["A1"] = "SUNRISE PRECISION SDN BHD"
    cover["A2"] = "Staff list for the attendance system"
    cover["A4"] = "Sheet 'Staff List' is the current one. 'Old List' is 2023 and stale."

    sheet = workbook.create_sheet("Staff List")
    sheet["A1"] = "SUNRISE PRECISION SDN BHD"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A2"] = "STAFF LIST — as at 01/08/2026"
    for column, header in HEADERS:
        cell = sheet.cell(row=4, column=column, value=header)
        cell.font = Font(bold=True)

    for index, row in enumerate(ROWS, start=1):
        number, remarks, name, dept, category, group, joined, left, device = row
        excel_row = 4 + index
        sheet.cell(row=excel_row, column=1, value=index)          # the decoy
        cell = sheet.cell(row=excel_row, column=2, value=number)  # text, zeros kept
        cell.number_format = "@"
        sheet.cell(row=excel_row, column=3, value=remarks)
        sheet.cell(row=excel_row, column=4, value=name)
        sheet.cell(row=excel_row, column=5, value=dept)
        sheet.cell(row=excel_row, column=6, value=category)
        sheet.cell(row=excel_row, column=7, value=group)
        sheet.cell(row=excel_row, column=9, value=joined)
        sheet.cell(row=excel_row, column=10, value=left)
        sheet.cell(row=excel_row, column=11, value=device)

    # A blank row and a footer, which is what real sheets have under the data.
    sheet.cell(row=4 + len(ROWS) + 2, column=2, value="")
    sheet.cell(row=4 + len(ROWS) + 3, column=4, value="Prepared by HR")

    stale = workbook.create_sheet("Old List")
    stale["A1"] = "2023 list — do not use"
    stale["A2"] = "Emp No"
    stale["B2"] = "Name"
    stale["A3"] = "0090"
    stale["B3"] = "Lim Wei Sheng"

    for column, width in (("A", 6), ("B", 10), ("C", 12), ("D", 28), ("E", 16),
                          ("F", 22), ("G", 14), ("H", 4), ("I", 14), ("J", 14),
                          ("K", 12)):
        sheet.column_dimensions[column].width = width
    return workbook


def main() -> int:
    target = Path(__file__).resolve().parent.parent / "fixtures" / "employees_sample.xlsx"
    target.parent.mkdir(parents=True, exist_ok=True)
    build().save(target)
    print(f"wrote {target}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
